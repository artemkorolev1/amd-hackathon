#!/usr/bin/env python3
"""Grade multi-model Excel results against ground truth from eval JSONs.

Uses the official fuzzy_match cascade from scripts/evaluate.py.
Outputs per-model accuracy + per-category breakdown.
"""

import json, os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from scripts.evaluate import fuzzy_match

import openpyxl

HELDOUT = "input/heldout_40.json"
DEV     = "input/dev_40.json"
RESULTS = [
    ("run_008 (heldout)", "eval_results/run_008_2026-07-12_221217.xlsx", "input/heldout_40.json"),
    ("run_009 (dev)",     "eval_results/run_009_2026-07-12_221403.xlsx", "input/dev_40.json"),
    ("run_001 (cx40)",    "eval_results/run_001_2026-07-12_222425.xlsx", "input/complexity_40.json"),
    ("run_002 (cx300)",   "eval_results/run_002_2026-07-12_222913.xlsx", "input/cx_300.json"),
]

def load_gold(path):
    with open(path) as f:
        data = json.load(f)
    questions = data if isinstance(data, list) else data.get("questions", data)
    gold = {}
    for q in questions:
        tid = q.get("task_id", "")
        gold_data = q.get("gold") or q.get("expected_answer", "")
        answer = ""
        if isinstance(gold_data, str):
            answer = gold_data
        elif isinstance(gold_data, dict):
            answer = gold_data.get("answer", "")
        gold[tid] = {
            "prompt": q.get("prompt", ""),
            "answer": answer,
            "category": q.get("category") or q.get("category_label", "unknown"),
            "difficulty": str(q.get("difficulty", q.get("complexity", q.get("complexity_score", "")))),
        }
    return gold

print("=" * 90)
print("  Multi-Model Evaluation — Official Fuzzy Match Grader")
print("=" * 90)

for run_label, xlsx_path, eval_path in RESULTS:
    gold = load_gold(eval_path)
    wb = openpyxl.load_workbook(xlsx_path)
    qs = wb["Questions"]

    # Group rows by model
    model_rows: dict[str, list] = {}
    for ri in range(2, qs.max_row + 1):
        model = qs.cell(ri, 3).value or "unknown"
        if model not in model_rows:
            model_rows[model] = []
        model_rows[model].append(ri)

    print(f"\n{'─'*90}")
    print(f"  {run_label} — {eval_path.split('/')[-1]} ({len(gold)} questions, {len(model_rows)} models)")
    print(f"{'─'*90}")

    for model_name in sorted(model_rows.keys(), key=str.lower):
        rows = model_rows[model_name]
        correct = 0
        total = 0
        cat_correct: dict[str, int] = {}
        cat_total: dict[str, int] = {}

        for ri in rows:
            tid = str(qs.cell(ri, 2).value or "")
            answer = str(qs.cell(ri, 5).value or "")
            expected_rec = gold.get(tid)
            if not expected_rec:
                continue
            expected = str(expected_rec["answer"]) if expected_rec["answer"] else ""
            category = expected_rec["category"]
            cat_total[category] = cat_total.get(category, 0) + 1
            total += 1
            if fuzzy_match(answer, expected):
                correct += 1
                cat_correct[category] = cat_correct.get(category, 0) + 1

        # Categories with text-based answer grading (no code execution needed)
        answer_cats = {"factual_knowledge", "logical_reasoning",
                       "math_reasoning", "sentiment_classification",
                       "factual", "logic", "math", "sentiment"}
        gradable = sum(cat_total.get(c, 0) for c in answer_cats)
        graded_correct = sum(cat_correct.get(c, 0) for c in answer_cats)

        # Difficulty breakdown
        diff_correct: dict[str, int] = {}
        diff_total: dict[str, int] = {}
        for ri in rows:
            tid = str(qs.cell(ri, 2).value or "")
            answer = str(qs.cell(ri, 5).value or "")
            expected_rec = gold.get(tid)
            if not expected_rec:
                continue
            difficulty = expected_rec.get("difficulty", "")
            cat = expected_rec.get("category", "")
            if cat not in answer_cats:
                continue
            diff_total[difficulty] = diff_total.get(difficulty, 0) + 1
            expected = str(expected_rec["answer"]) if expected_rec["answer"] else ""
            if fuzzy_match(answer, expected):
                diff_correct[difficulty] = diff_correct.get(difficulty, 0) + 1

        pct = (correct / total * 100) if total else 0
        gradable_pct = (graded_correct / gradable * 100) if gradable else 0
        print(f"\n  {model_name}")
        print(f"    Overall (all {total} q):     {correct}/{total} = {pct:.1f}%  {'✅' if pct >= 84.2 else '❌'}")
        if gradable:
            print(f"    Text-answer ({gradable} q):  {graded_correct}/{gradable} = {gradable_pct:.1f}%")
        if diff_total:
            parts = []
            for d in ["simple", "medium", "hard"]:
                if d in diff_total:
                    c = diff_correct.get(d, 0)
                    t = diff_total[d]
                    parts.append(f"{d}: {c}/{t} = {c/t*100:.0f}%")
            print(f"    By difficulty:            {' | '.join(parts)}")
        for cat in sorted(cat_total.keys()):
            c = cat_correct.get(cat, 0)
            t = cat_total[cat]
            cpct = (c / t * 100) if t else 0
            print(f"      {cat:25s}  {c:2d}/{t:2d} = {cpct:5.1f}%")

wb.close()
