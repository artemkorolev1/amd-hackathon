#!/usr/bin/env python3
"""Generate step-by-step Excel report from the V12E pipeline trace."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Pipeline Trace"

# Column headers
HEADERS = ["Step", "Module", "Latency (s)", "Input", "Output"]
for i, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Data from the successful test run
steps = [
    (1, "pre_filter (was stage0)", 0.0000,
     "What is 2+2?",
     "action=bypass, category=math_arithmetic"),
    (2, "category_filter (was stage2)", 0.0057,
     "What is 2+2?", 
     "category=math, confidence=0.9000, top3={math:6.0, factual:2.0, logic:0.0}"),
    (3, "complexity_filter (was stage3)", 0.0015,
     "category=math",
     "score=0.1106 → low complexity"),
    (4, "deterministic (solve_arithmetic)", 0.0055,
     "category=math_arithmetic",
     "answer=4"),
    (5, "local_model (was lora_model)", 0.0000,
     "— (bypassed by deterministic)",
     "— not invoked"),
]

for row_idx, (step, module, lat, inp, out) in enumerate(steps, 2):
    ws.cell(row=row_idx, column=1, value=step).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=2, value=module)
    ws.cell(row=row_idx, column=3, value=lat)
    ws.cell(row=row_idx, column=3).number_format = "0.0000"
    ws.cell(row=row_idx, column=4, value=inp[:80] + ("..." if len(inp) > 80 else ""))
    ws.cell(row=row_idx, column=5, value=out[:120] if out else "")

# Add a second sheet: Rename mapping
ws2 = wb.create_sheet("Module Renames")
ws2_headers = ["Old Name", "New Name", "Exports"]
for i, h in enumerate(ws2_headers, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

renames = [
    ("agent/stage0.py", "agent/pre_filter.py", "stage0(), Stage0Result, RE_* regexes"),
    ("agent/stage2.py", "agent/category_filter.py", "classify(), CATEGORIES_8WAY, PRIORITY"),
    ("agent/stage3.py", "agent/complexity_filter.py", "score(), describe(), validate()"),
    ("agent/solvers/lora_model.py", "agent/solvers/local_model.py", "chat_completion(), _get_model()"),
]
for i, (old, new, exports) in enumerate(renames, 2):
    ws2.cell(row=i, column=1, value=old)
    ws2.cell(row=i, column=2, value=new)
    ws2.cell(row=i, column=3, value=exports)

# Add a third sheet: Fixes Applied
ws3 = wb.create_sheet("Code Fixes")
ws3_headers = ["File", "Fix"]
for i, h in enumerate(ws3_headers, 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

fixes = [
    ("run_v12e.py", "DET_CAT_MAP: removed code_gen→code_debugging, factual→factual_knowledge"),
    ("eval_v12e.py", "DET_CAT_MAP: same fix + removed DETERMINISTIC_CATEGORIES dead set"),
    ("run_v12e.py, eval_v12e.py", "except: pass → except Exception: logger.debug(...)"),
    ("eval_v12e.py", "Added few-shot NER/Sentiment/Math examples (was missing)"),
    ("eval_v12e.py", "Removed dead use_lora=True variable"),
    ("eval_v12e.py", "Renamed answer_source labels: lora_* → local_*"),
    ("run_v12e.py, eval_v12e.py", "Removed stale LORA_ADAPTERS_DIR env var"),
    ("agent/main.py", "Updated all import paths to new module names"),
    ("agent/solvers/local_model.py", "Stripped LoRA adapter code (adapters don't exist)"),
    ("eval_results/", "Deleted 20 stale eval files (kept 4 latest)"),
    ("archived/lora/", "Moved LoRA-specific files (training data, old sources)"),
]
for i, (file, fix) in enumerate(fixes, 2):
    ws3.cell(row=i, column=1, value=file)
    ws3.cell(row=i, column=2, value=fix)

# Adjust column widths
for ws_sheet in [ws, ws2, ws3]:
    for col in ws_sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws_sheet.column_dimensions[col[0].column_letter].width = min(max_len, 60)

outpath = "/home/artem/dev/amd-hackathon/eval_results/v12e_rename_report.xlsx"
os.makedirs(os.path.dirname(outpath), exist_ok=True)
wb.save(outpath)
print(f"Saved: {outpath}")
