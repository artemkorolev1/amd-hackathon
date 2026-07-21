#!/usr/bin/env python3
"""
scripts/eval_framework.py — AMD Hackathon Evaluation Framework.

Orchestrates the full pipeline run against an eval set:
  1. Accepts eval set path + mode (gpu/cpu, default gpu)
  2. Normalises the eval set (auto-adds task_ids where missing)
  3. Runs harness.py as a subprocess (full pipeline)
  4. Extracts per-question timings from RunLogger Excel
  5. Grades results using runner/evaluate.py grade_results()
  6. Produces structured JSON → eval_results/eval_{timestamp}.json
  7. Produces XLSX report  → eval_results/eval_report_{timestamp}.xlsx
  8. Prints comprehensive summary (total, passed, failed, accuracy, timing)

Usage:
    python scripts/eval_framework.py data/eval/primary/eval_mini_10.json --gpu
    python scripts/eval_framework.py data/eval/primary/eval_hard_218.json
    python scripts/eval_framework.py data/eval/validation-v3.json --cpu --verbose
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import subprocess
import sys
import tempfile
import time
from datetime import datetime

# ── Paths ──────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_HERE)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

_EVAL_RESULTS_DIR = os.path.join(_PROJECT_ROOT, "eval_results")
_HARNESS_PATH = os.path.join(_HERE, "harness.py")
_HARNESS_XLSX_DIR = os.path.join(_HERE, "eval_results")

logger = logging.getLogger("eval_framework")


# ═════════════════════════════════════════════════════════════════════════════
# Helpers
# ═════════════════════════════════════════════════════════════════════════════


def _normalise_eval_set(eval_path: str) -> str:
    """Ensure every entry in the eval set has a *task_id*.

    The harness generates ``idx_{i}`` for entries missing a ``task_id``, but
    ``runner/evaluate.py`` silently skips entries without one.  We inject
    matching ``idx_{i}`` keys into a temporary normalised copy so the grader
    sees the same IDs the pipeline produced.

    Returns the *same path* if no normalisation was needed, otherwise a path
    to a temporary file that the caller **must** delete.
    """
    with open(eval_path) as f:
        raw = json.load(f)

    if isinstance(raw, list):
        entries: list[dict] = raw
        fmt_is_list = True
    elif isinstance(raw, dict):
        entries = raw.get("questions", raw.get("tasks", []))
        fmt_is_list = False
    else:
        raise ValueError(f"Unsupported JSON root type: {type(raw).__name__}")

    changed = False
    normalised: list[dict] = []
    for i, entry in enumerate(entries):
        entry = dict(entry)  # shallow copy
        tid = entry.get("task_id")
        if not tid:
            entry["task_id"] = f"idx_{i}"
            changed = True
        normalised.append(entry)

    if not changed:
        return eval_path

    tmp = tempfile.NamedTemporaryFile(
        mode="w",
        suffix=".json",
        prefix="eval_norm_",
        delete=False,
    )
    if fmt_is_list:
        json.dump(normalised, tmp)
    else:
        out = dict(raw)
        if "questions" in out:
            out["questions"] = normalised
        else:
            out["tasks"] = normalised
        json.dump(out, tmp)
    tmp.close()

    added = sum(1 for e in normalised if e["task_id"].startswith("idx_"))
    logger.info("Normalised eval set: added task_id to %d entries", added)
    return tmp.name


# ── RunLogger XLSX helpers ─────────────────────────────────────────────────


def _find_latest_runlogger_xlsx(xlsx_dir: str) -> str | None:
    """Return the most recent ``run_*.xlsx`` in *xlsx_dir*, or *None*."""
    import glob

    files = glob.glob(os.path.join(xlsx_dir, "run_*.xlsx"))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def _extract_timings_from_runlogger(xlsx_path: str) -> dict[str, float]:
    """Parse per-task timing from the RunLogger's *Questions* sheet.

    Returns a dict mapping ``task_id → total_ms``.
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Questions" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Questions"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    try:
        idx_tid = headers.index("Task ID")
        idx_ms = headers.index("Total (ms)")
    except ValueError:
        wb.close()
        return {}

    timings: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = str(row[idx_tid]) if row[idx_tid] is not None else ""
        t_ms = row[idx_ms]
        if tid and t_ms is not None:
            timings[tid] = float(t_ms)

    wb.close()
    return timings


# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════


def main() -> int:
    parser = argparse.ArgumentParser(
        description="AMD Hackathon Evaluation Framework — run, grade, report.",
    )
    parser.add_argument("eval_path", help="Path to eval set JSON file")
    parser.add_argument(
        "--gpu",
        action="store_true",
        help="GPU mode: all layers on GPU (DEFAULT)",
    )
    parser.add_argument(
        "--cpu",
        action="store_true",
        help="CPU mode: zero layers offloaded",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Show verbose logging",
    )
    args = parser.parse_args()

    # ── 0. Mode (default: gpu) ─────────────────────────────────────────────
    mode = "gpu"
    if args.cpu:
        mode = "cpu"
    elif args.gpu:
        mode = "gpu"

    flag = "--gpu" if mode == "gpu" else "--cpu"
    os.environ["N_GPU_LAYERS"] = "-1" if mode == "gpu" else "0"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(_EVAL_RESULTS_DIR, exist_ok=True)
    os.makedirs(_HARNESS_XLSX_DIR, exist_ok=True)

    # ── Header ──────────────────────────────────────────────────────────────
    print()
    print("=" * 60)
    print("  AMD HACKATHON — EVALUATION FRAMEWORK")
    print("=" * 60)
    print(f"  Mode:          {mode.upper()}  "
          f"(N_GPU_LAYERS={os.environ['N_GPU_LAYERS']})")
    print(f"  Eval set:      {os.path.abspath(args.eval_path)}")
    print(f"  Timestamp:     {timestamp}")
    print("=" * 60)
    print()

    normalised_path = args.eval_path
    enriched_path: str | None = None
    exit_code = 1

    try:
        # ── 1. Normalise eval set ──────────────────────────────────────────
        print("[1/5] Preparing eval set\u2026")
        normalised_path = _normalise_eval_set(args.eval_path)
        if normalised_path != args.eval_path:
            print("  \u2713 Added task_ids to entries lacking them")
        print(f"  \u2713 Eval set: {os.path.basename(normalised_path)}")

        # ── 2. Run the pipeline via harness.py ─────────────────────────────
        print()
        print("[2/5] Running full pipeline via harness.py\u2026")
        t_start = time.monotonic()

        pipe_result = subprocess.run(
            [sys.executable, "-m", "scripts.harness", flag, normalised_path],
            capture_output=True,
            text=True,
            timeout=1800,
        )

        t_pipeline = time.monotonic() - t_start

        if pipe_result.returncode != 0:
            print(f"  \u26a0 Harness exited with code {pipe_result.returncode}")
            if pipe_result.stderr:
                for line in pipe_result.stderr.strip().splitlines()[-5:]:
                    print(f"  \u26a0 {line}")
        else:
            print(f"  \u2713 Pipeline complete ({t_pipeline:.1f}s)")

        if args.verbose and pipe_result.stdout:
            for line in pipe_result.stdout.strip().splitlines()[:10]:
                print(f"    {line[:120]}")

        # ── 3. Read predictions ────────────────────────────────────────────
        print()
        print("[3/5] Reading predictions & grading\u2026")

        results_path = "/output/results.json"
        alt_results_path = os.path.join(_PROJECT_ROOT, "output", "results.json")

        if not os.path.exists(results_path) and os.path.exists(alt_results_path):
            results_path = alt_results_path

        if not os.path.exists(results_path):
            print(f"  \u2717 No predictions found "
                  f"(checked {results_path})")
            return 1

        with open(results_path) as f:
            predictions_raw = json.load(f)

        pred_entries: list[dict] = (
            predictions_raw
            if isinstance(predictions_raw, list)
            else predictions_raw.get("results", [])
        )
        print(f"  \u2713 Loaded {len(pred_entries)} predictions")

        # ── 4. Extract per-question timings ─────────────────────────────────
        latest_xlsx = _find_latest_runlogger_xlsx(_HARNESS_XLSX_DIR)
        timings_map: dict[str, float] = {}
        if latest_xlsx:
            timings_map = _extract_timings_from_runlogger(latest_xlsx)
            print(f"  \u2713 Extracted timings for {len(timings_map)} tasks "
                  f"from RunLogger")

        for entry in pred_entries:
            tid = entry.get("task_id", "")
            entry["timing_ms"] = timings_map.get(tid, 0.0)

        # Stash enriched predictions for grade_results
        fd, enriched_path = tempfile.mkstemp(
            suffix=".json",
            prefix="enriched_",
            dir=_EVAL_RESULTS_DIR,
            text=True,
        )
        with os.fdopen(fd, "w") as f:
            json.dump(pred_entries, f)

        # ── 5. Grade & produce XLSX report via runner.evaluate ─────────────
        from runner.evaluate import grade_results

        xlsx_path = os.path.join(
            _EVAL_RESULTS_DIR, f"eval_report_{timestamp}.xlsx"
        )

        report = grade_results(
            results_json=enriched_path,
            gold_json=normalised_path,
            output_xlsx=xlsx_path,
            verbose=args.verbose,
        )
        print(f"  \u2713 XLSX report: {xlsx_path}")

        # ── 6. Write structured JSON ───────────────────────────────────────
        print()
        print("[4/5] Writing structured JSON report\u2026")

        report["run_info"] = {
            "mode": mode,
            "n_gpu_layers": int(os.environ["N_GPU_LAYERS"]),
            "eval_path": os.path.abspath(args.eval_path),
            "timestamp": timestamp,
            "pipeline_elapsed_s": round(t_pipeline, 2),
            "harness_returncode": pipe_result.returncode,
        }

        json_path = os.path.join(
            _EVAL_RESULTS_DIR, f"eval_{timestamp}.json"
        )
        with open(json_path, "w") as f:
            json.dump(report, f, indent=2, default=str)
        print(f"  \u2713 JSON report:  {json_path}")

        # ── 7. Print summary ───────────────────────────────────────────────
        print()
        print("[5/5] Summary")
        print()
        print("=" * 60)
        print("  EVALUATION SUMMARY")
        print("=" * 60)

        overall = report["overall"]
        print(f"  Mode:          {mode.upper()}")
        print(f"  Total tasks:   {overall['total']}")
        print(f"  Passed:        {overall['correct']}")
        print(f"  Failed:        {overall['total'] - overall['correct']}")
        print(f"  Accuracy:      {overall['accuracy']:.1%}")
        gate_str = "\u2713 PASSED" if overall["gate_pass"] else "\u2717 FAILED"
        print(f"  84.2% Gate:    {gate_str}")

        if report.get("by_category"):
            print()
            print(f"  {'Category':30s}  {'Score':>10s}  {'Accuracy':>8s}")
            dashline = "-" * 30
            dash10 = "-" * 10
            dash8 = "-" * 8
            print(f"  {dashline}  {dash10:>10s}  {dash8:>8s}")
            for cat, stats in sorted(report["by_category"].items()):
                acc = stats["accuracy"]
                score = f"{stats['correct']}/{stats['total']}"
                print(f"  {cat:30s}  {score:>10s}  {acc:>7.1%}")

        timing = report.get("timing", {})
        if timing.get("mean", 0) > 0:
            print()
            print(f"  Timing (ms)")
            print(f"    Mean:   {timing['mean']:.1f}")
            print(f"    Median: {timing['median']:.1f}")
            print(f"    P95:    {timing['p95']:.1f}")

        print(f"  Wall time:      {t_pipeline:.1f}s")
        print("=" * 60)
        print()

        exit_code = 0 if overall["gate_pass"] else 0  # always exit 0 for now

    finally:
        # ── Cleanup temp files ────────────────────────────────────────────
        if enriched_path is not None:
            try:
                os.remove(enriched_path)
            except OSError:
                pass
        if normalised_path != args.eval_path:
            try:
                os.remove(normalised_path)
            except OSError:
                pass

    return exit_code


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
    )
    sys.exit(main())
