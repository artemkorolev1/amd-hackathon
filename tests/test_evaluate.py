#!/usr/bin/env python3
"""Tests for runner/evaluate.py - grading pipeline against ground truth."""

import json
import os
import sys
import tempfile
import unittest
from unittest.mock import patch, MagicMock, mock_open

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from scripts.evaluate import fuzzy_match, grade_answer


# =========================================================================
# Tests for fuzzy_match (reused from scripts/evaluate)
# =========================================================================

class TestFuzzyMatch(unittest.TestCase):
    """Verify the fuzzy_match cascade directly."""

    def test_exact_case_insensitive(self):
        assert fuzzy_match("Canberra", "Canberra") is True
        assert fuzzy_match("canberra", "Canberra") is True
        assert fuzzy_match("CANBERRA", "Canberra") is True

    def test_substring_expected_in_answer(self):
        assert fuzzy_match("The capital is Canberra", "Canberra") is True

    def test_substring_short_answer_in_expected(self):
        assert fuzzy_match("yes", "yes, that is correct") is True

    def test_substring_too_short(self):
        # len(a) < 3, so short answer in expected is not checked
        assert fuzzy_match("no", "nope") is False

    def test_numeric_tolerance_pairwise(self):
        assert fuzzy_match("6.0", "6") is True
        # Note: "6" is a substring of "6.05" so substring check catches it
        assert fuzzy_match("6.05", "6") is True
        # Use non-substring numbers for >1% check
        assert fuzzy_match("7.0", "6.0") is False    # 16.7% > 1%

    def test_numeric_single_number_in_answer(self):
        assert fuzzy_match("The answer is 72 km/h", "72") is True
        # "72" is not a substring of "73", and 73 is 1.4% off from 72
        assert fuzzy_match("The answer is 73", "72") is False  # 1.4% > 1%

    def test_token_overlap_short_answer(self):
        # Short expected (< 50 chars): need >= 50% token overlap
        assert fuzzy_match("Gabriel Garcia Marquez wrote it", "Gabriel Garcia Marquez") is True

    def test_token_overlap_long_answer(self):
        # Long expected (>= 50 chars): need >= 30% token overlap
        expected = "it did not rain because the ground is not wet"
        answer = "it did not rain"
        assert fuzzy_match(answer, expected) is True

    def test_no_match(self):
        assert fuzzy_match("Paris", "London") is False

    def test_empty_expected(self):
        assert fuzzy_match("anything", "") is True
        assert fuzzy_match("", "") is False

    def test_numeric_zero_tolerance(self):
        # "0" is a substring of "0.005", so this passes via substring check
        assert fuzzy_match("0.005", "0") is True
        # Use non-substring numbers: 0.02 vs 0.0 — "0" is still substring of "0.02",
        # so for a proper >0.01 mismatch we test both values that don't overlap
        assert fuzzy_match("200", "100") is False   # 100% > 1%
        # But "0.005" is within 0.01 absolute of 0 via numeric check too
        assert fuzzy_match("0.005", "0") is True

    def test_numeric_negative(self):
        assert fuzzy_match("-5", "-5") is True
        assert fuzzy_match("-5.05", "-5") is True   # within 1%


# =========================================================================
# Tests for grade_answer
# =========================================================================

class TestGradeAnswer(unittest.TestCase):
    """Verify grade_answer handles various formats."""

    def test_exact_pass(self):
        passed, reason = grade_answer("Canberra", "Canberra")
        assert passed is True
        assert reason == "Passed"

    def test_empty_answer(self):
        passed, reason = grade_answer("", "something")
        assert passed is False
        assert "Empty answer" in reason

    def test_agent_error(self):
        passed, reason = grade_answer("[ERROR] API timeout", "something")
        assert passed is False
        assert "Agent error" in reason

    def test_numeric_as_string(self):
        passed, reason = grade_answer("6", "6")
        assert passed is True

    def test_fuzzy_match_fallback(self):
        passed, reason = grade_answer("The answer is Canberra", "Canberra")
        assert passed is True

    def test_failure_diagnostic(self):
        passed, reason = grade_answer("Paris", "London")
        assert passed is False
        assert "London" in reason or "expected" in reason


# =========================================================================
# Helper to create temporary JSON files for load tests
# =========================================================================

def _make_temp_json(data):
    """Write data to a temp JSON file, return the path."""
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False)
    json.dump(data, tmp)
    tmp.close()
    return tmp.name


# =========================================================================
# Tests for load_gold
# =========================================================================

class TestLoadGold(unittest.TestCase):
    """Verify load_gold handles both gold formats."""

    def test_dev40_format(self):
        """dev_40.json format: {gold: {answer: ..., accept?: [...]}}."""
        from runner.evaluate import load_gold
        data = [
            {"task_id": "fact-1", "category": "factual_knowledge",
             "prompt": "What is the capital?",
             "gold": {"answer": "Canberra"}},
            {"task_id": "math-1", "category": "math_reasoning",
             "prompt": "What is 15% of 240?",
             "gold": {"answer": 36.0}},
            {"task_id": "ner-1", "category": "ner",
             "prompt": "Extract entities.",
             "gold": {"entities": ["Tim", "Apple"]}},
        ]
        path = _make_temp_json(data)
        try:
            result = load_gold(path)
            assert "fact-1" in result
            assert result["fact-1"]["expected"] == "Canberra"
            assert result["fact-1"]["category"] == "factual_knowledge"
            assert result["fact-1"]["prompt"] == "What is the capital?"

            assert "math-1" in result
            assert result["math-1"]["expected"] == "36.0"  # numeric → str

            # NER entities: joined into expected string
            assert "ner-1" in result
            assert "Tim" in result["ner-1"]["expected"]
            assert "Apple" in result["ner-1"]["expected"]
        finally:
            os.unlink(path)

    def test_two_questions_format(self):
        """two_questions.json format: {expected_answer, difficulty}."""
        from runner.evaluate import load_gold
        data = [
            {"task_id": "tc8", "category": "factual",
             "prompt": "What is the capital of France?",
             "expected_answer": "Paris",
             "difficulty": "simple"},
            {"task_id": "tc7", "category": "math",
             "prompt": "What is 25% of 360?",
             "expected_answer": "90",
             "difficulty": "simple"},
        ]
        path = _make_temp_json(data)
        try:
            result = load_gold(path)
            assert "tc8" in result
            assert result["tc8"]["expected"] == "Paris"
            assert result["tc8"]["category"] == "factual"
            assert result["tc8"]["difficulty"] == "simple"
            assert "tc7" in result
            assert result["tc7"]["expected"] == "90"
        finally:
            os.unlink(path)

    def test_gold_with_accept_list(self):
        """gold.accept list → aliases included in expected."""
        from runner.evaluate import load_gold
        data = [
            {"task_id": "fact-4", "category": "factual_knowledge",
             "prompt": "Who wrote 100 Years?",
             "gold": {"answer": "Gabriel Garcia Marquez",
                      "accept": ["Gabriel Garcia Marquez", "Garcia Marquez"]}},
        ]
        path = _make_temp_json(data)
        try:
            result = load_gold(path)
            assert result["fact-4"]["expected"] == "Gabriel Garcia Marquez"
        finally:
            os.unlink(path)

    def test_gold_with_code_task(self):
        """Code tasks with function/tests should still be loaded."""
        from runner.evaluate import load_gold
        data = [
            {"task_id": "debug-1", "category": "code_debugging",
             "prompt": "Fix the bug.",
             "gold": {"function": "average", "tests": []}},
        ]
        path = _make_temp_json(data)
        try:
            result = load_gold(path)
            assert "debug-1" in result
            assert result["debug-1"]["expected"] == ""  # no string answer
        finally:
            os.unlink(path)

    def test_missing_task_id(self):
        """Task without task_id gets skipped."""
        from runner.evaluate import load_gold
        data = [
            {"category": "math", "prompt": "2+2?",
             "gold": {"answer": "4"}},
        ]
        path = _make_temp_json(data)
        try:
            result = load_gold(path)
            assert len(result) == 0
        finally:
            os.unlink(path)


# =========================================================================
# Tests for load_predictions
# =========================================================================

class TestLoadPredictions(unittest.TestCase):
    """Verify load_predictions parses BatchRunner output."""

    def test_standard_format(self):
        from runner.evaluate import load_predictions
        data = [
            {"task_id": "fact-1", "answer": "Canberra", "timing_ms": 123.4},
            {"task_id": "math-1", "answer": "36", "timing_ms": 45.6},
        ]
        path = _make_temp_json(data)
        try:
            result = load_predictions(path)
            assert "fact-1" in result
            assert result["fact-1"]["answer"] == "Canberra"
            assert result["fact-1"]["timing_ms"] == 123.4
            assert "math-1" in result
            assert result["math-1"]["answer"] == "36"
        finally:
            os.unlink(path)

    def test_with_additional_fields(self):
        from runner.evaluate import load_predictions
        data = [
            {"task_id": "t1", "answer": "yes", "timing_ms": 100,
             "worker": 0, "extra": "ignored"},
        ]
        path = _make_temp_json(data)
        try:
            result = load_predictions(path)
            assert result["t1"]["answer"] == "yes"
            assert result["t1"]["timing_ms"] == 100
        finally:
            os.unlink(path)

    def test_empty_predictions(self):
        from runner.evaluate import load_predictions
        path = _make_temp_json([])
        try:
            result = load_predictions(path)
            assert result == {}
        finally:
            os.unlink(path)

    def test_no_timing_field(self):
        from runner.evaluate import load_predictions
        data = [
            {"task_id": "t1", "answer": "Paris"},
        ]
        path = _make_temp_json(data)
        try:
            result = load_predictions(path)
            assert result["t1"]["answer"] == "Paris"
            assert result["t1"]["timing_ms"] == 0.0
        finally:
            os.unlink(path)


# =========================================================================
# Tests for evaluate_tasks
# =========================================================================

class TestEvaluateTasks(unittest.TestCase):
    """Verify grading logic."""

    def test_all_correct(self):
        from runner.evaluate import evaluate_tasks
        gold = {
            "t1": {"prompt": "Capital of France?", "expected": "Paris",
                   "category": "factual", "difficulty": "simple"},
            "t2": {"prompt": "2+2?", "expected": "4",
                   "category": "math", "difficulty": "simple"},
        }
        preds = {
            "t1": {"answer": "Paris", "timing_ms": 10.0},
            "t2": {"answer": "4", "timing_ms": 5.0},
        }
        results = evaluate_tasks(gold, preds)
        assert len(results) == 2
        assert all(r["correct"] for r in results)

    def test_some_wrong(self):
        from runner.evaluate import evaluate_tasks
        gold = {
            "t1": {"prompt": "Capital?", "expected": "Paris", "category": "factual"},
            "t2": {"prompt": "2+2?", "expected": "4", "category": "math"},
        }
        preds = {
            "t1": {"answer": "London", "timing_ms": 10.0},
            "t2": {"answer": "4", "timing_ms": 5.0},
        }
        results = evaluate_tasks(gold, preds)
        assert results[0]["correct"] is False
        assert results[0]["task_id"] == "t1"
        assert results[1]["correct"] is True
        assert results[1]["task_id"] == "t2"

    def test_missing_prediction(self):
        from runner.evaluate import evaluate_tasks
        gold = {
            "t1": {"prompt": "Capital?", "expected": "Paris", "category": "factual"},
            "t2": {"prompt": "2+2?", "expected": "4", "category": "math"},
        }
        preds = {
            "t1": {"answer": "Paris", "timing_ms": 10.0},
            # t2 missing
        }
        results = evaluate_tasks(gold, preds)
        assert results[0]["correct"] is True
        assert results[1]["correct"] is False
        assert results[1]["answer"] == ""

    def test_difficulty_preserved(self):
        from runner.evaluate import evaluate_tasks
        gold = {
            "t1": {"prompt": "Capital?", "expected": "Paris",
                   "category": "factual", "difficulty": "hard"},
        }
        preds = {
            "t1": {"answer": "Paris", "timing_ms": 10.0},
        }
        results = evaluate_tasks(gold, preds)
        assert results[0]["difficulty"] == "hard"
        assert results[0]["category"] == "factual"

    def test_with_accept_aliases(self):
        """Tasks with accept list should match aliases."""
        from runner.evaluate import evaluate_tasks
        gold = {
            "t1": {"prompt": "Who?", "expected": "Gabriel Garcia Marquez",
                   "category": "factual", "accept": ["Garcia Marquez"]},
        }
        preds = {
            "t1": {"answer": "Garcia Marquez", "timing_ms": 10.0},
        }
        results = evaluate_tasks(gold, preds)
        assert results[0]["correct"] is True

    def test_keyword_summarization(self):
        """Tasks with keywords should check coverage."""
        from runner.evaluate import evaluate_tasks
        gold = {
            "t1": {"prompt": "Summarize.", "expected": "",
                   "category": "summarization",
                   "keywords": ["Panama", "Canal", "shipping"],
                   "min_coverage": 0.5},
        }
        preds = {
            "t1": {"answer": "The Panama Canal connects oceans and helps shipping.",
                   "timing_ms": 10.0},
        }
        results = evaluate_tasks(gold, preds)
        assert results[0]["correct"] is True

    def test_keyword_summarization_fail(self):
        """Tasks with keywords where coverage is too low."""
        from runner.evaluate import evaluate_tasks
        gold = {
            "t1": {"prompt": "Summarize.", "expected": "",
                   "category": "summarization",
                   "keywords": ["Panama", "Canal", "shipping", "Atlantic"],
                   "min_coverage": 0.75},
        }
        preds = {
            "t1": {"answer": "It's about a canal.", "timing_ms": 10.0},
        }
        results = evaluate_tasks(gold, preds)
        assert results[0]["correct"] is False

    def test_code_task_skipped(self):
        """Code tasks (function/tests) should be marked specially."""
        from runner.evaluate import evaluate_tasks
        gold = {
            "t1": {"prompt": "Write a function.", "expected": "",
                   "category": "code_generation",
                   "function": "fizzbuzz", "tests": []},
        }
        preds = {
            "t1": {"answer": "def fizzbuzz...", "timing_ms": 10.0},
        }
        results = evaluate_tasks(gold, preds)
        # Code tasks pass by convention (not checked via fuzzy_match)
        assert results[0]["correct"] is True
        assert "code" in results[0]["reason"].lower()


# =========================================================================
# Tests for build_report
# =========================================================================

class TestBuildReport(unittest.TestCase):
    """Verify aggregation math."""

    def test_all_correct(self):
        from runner.evaluate import build_report
        results = [
            {"task_id": "t1", "category": "factual", "difficulty": "simple",
             "correct": True, "timing_ms": 10.0},
            {"task_id": "t2", "category": "math", "difficulty": "simple",
             "correct": True, "timing_ms": 20.0},
        ]
        report = build_report(results)
        assert report["overall"]["total"] == 2
        assert report["overall"]["correct"] == 2
        assert report["overall"]["accuracy"] == 1.0
        assert report["overall"]["gate_pass"] is True
        assert report["by_category"]["factual"]["total"] == 1
        assert report["by_category"]["factual"]["correct"] == 1
        assert report["by_category"]["math"]["total"] == 1
        assert report["by_category"]["math"]["correct"] == 1

    def test_mixed_accuracy_gate_fail(self):
        from runner.evaluate import build_report
        results = [
            {"task_id": "t1", "category": "factual", "difficulty": "simple",
             "correct": True, "timing_ms": 10.0},
            {"task_id": "t2", "category": "factual", "difficulty": "simple",
             "correct": False, "timing_ms": 20.0},
        ]
        report = build_report(results)
        assert report["overall"]["total"] == 2
        assert report["overall"]["correct"] == 1
        assert report["overall"]["accuracy"] == 0.5
        assert report["overall"]["gate_pass"] is False

    def test_by_difficulty(self):
        from runner.evaluate import build_report
        results = [
            {"task_id": "t1", "category": "factual", "difficulty": "simple",
             "correct": True, "timing_ms": 10.0},
            {"task_id": "t2", "category": "math", "difficulty": "hard",
             "correct": False, "timing_ms": 20.0},
            {"task_id": "t3", "category": "math", "difficulty": "hard",
             "correct": True, "timing_ms": 30.0},
        ]
        report = build_report(results)
        assert "simple" in report["by_difficulty"]
        assert report["by_difficulty"]["simple"]["total"] == 1
        assert report["by_difficulty"]["simple"]["correct"] == 1
        assert report["by_difficulty"]["hard"]["total"] == 2
        assert report["by_difficulty"]["hard"]["correct"] == 1

    def test_failures_list(self):
        from runner.evaluate import build_report
        results = [
            {"task_id": "t1", "category": "factual", "correct": True,
             "reason": "Passed", "timing_ms": 10.0},
            {"task_id": "t2", "category": "math", "correct": False,
             "reason": "Wrong", "timing_ms": 20.0},
        ]
        report = build_report(results)
        assert len(report["failures"]) == 1
        assert report["failures"][0]["task_id"] == "t2"

    def test_timing_stats(self):
        from runner.evaluate import build_report
        results = [
            {"task_id": "t1", "category": "factual", "correct": True,
             "timing_ms": 10.0},
            {"task_id": "t2", "category": "math", "correct": True,
             "timing_ms": 20.0},
            {"task_id": "t3", "category": "math", "correct": False,
             "timing_ms": 90.0},
        ]
        report = build_report(results)
        assert report["timing"]["mean"] == 40.0
        assert report["timing"]["median"] == 20.0
        assert "math" in report["timing"]["per_category"]
        assert "factual" in report["timing"]["per_category"]

    def test_empty_results(self):
        from runner.evaluate import build_report
        report = build_report([])
        assert report["overall"]["total"] == 0
        assert report["overall"]["accuracy"] == 0.0
        assert report["overall"]["gate_pass"] is False

    def test_per_task_in_report(self):
        from runner.evaluate import build_report
        results = [
            {"task_id": "t1", "category": "factual", "difficulty": "simple",
             "prompt": "Q?", "expected": "A", "answer": "A",
             "correct": True, "reason": "Passed", "timing_ms": 10.0},
        ]
        report = build_report(results)
        assert len(report["per_task"]) == 1
        assert report["per_task"][0]["task_id"] == "t1"

    def test_accuracy_at_gate_edge(self):
        """84.2% gate boundary: 5/6 = 83.3% fails, 6/7 = 85.7% passes."""
        from runner.evaluate import build_report
        # 5 correct out of 6 = 83.33% → FAIL
        results = [{"task_id": f"t{i}", "category": "factual",
                     "correct": i < 5, "timing_ms": 10.0}
                    for i in range(6)]
        report = build_report(results)
        assert report["overall"]["gate_pass"] is False

        # 6 correct out of 7 = 85.71% → PASS
        results2 = [{"task_id": f"t{i}", "category": "factual",
                      "correct": i < 6, "timing_ms": 10.0}
                     for i in range(7)]
        report2 = build_report(results2)
        assert report2["overall"]["gate_pass"] is True


# =========================================================================
# Tests for write_xlsx
# =========================================================================

class TestWriteXLSX(unittest.TestCase):
    """Verify XLSX file creation (integration-level)."""

    def test_basic_xlsx_creation(self):
        from runner.evaluate import build_report, write_xlsx
        results = [
            {"task_id": "t1", "category": "factual", "difficulty": "simple",
             "prompt": "Capital?", "expected": "Paris", "answer": "Paris",
             "correct": True, "reason": "Passed", "timing_ms": 10.0},
            {"task_id": "t2", "category": "math", "difficulty": "hard",
             "prompt": "2+2?", "expected": "4", "answer": "5",
             "correct": False, "reason": "Wrong", "timing_ms": 20.0},
        ]
        report = build_report(results)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_xlsx(report, tmp.name)
            # Verify the file was created and is non-empty
            assert os.path.exists(tmp.name)
            assert os.path.getsize(tmp.name) > 0

            # Verify it's a valid XLSX with 3 sheets
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            assert wb.sheetnames == ["Summary", "Details", "Failures"]
            # Check Summary sheet has content
            summary = wb["Summary"]
            assert summary.cell(1, 1).value is not None
            # Check Details sheet has headers + data
            details = wb["Details"]
            assert details.cell(1, 1).value == "Task ID"
            assert details.max_row == 3  # header + 2 rows
            # Check Failures sheet has 1 data row + header
            failures = wb["Failures"]
            assert failures.cell(1, 1).value == "Task ID"
            assert failures.max_row == 2  # header + 1 failure
            wb.close()
        finally:
            os.unlink(tmp.name)

    def test_no_failures_sheet_still_created(self):
        from runner.evaluate import build_report, write_xlsx
        results = [
            {"task_id": "t1", "category": "factual", "difficulty": "simple",
             "prompt": "Capital?", "expected": "Paris", "answer": "Paris",
             "correct": True, "reason": "Passed", "timing_ms": 10.0},
        ]
        report = build_report(results)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            write_xlsx(report, tmp.name)
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            assert "Failures" in wb.sheetnames
            failures = wb["Failures"]
            # Only header row since no failures
            assert failures.max_row == 1
            wb.close()
        finally:
            os.unlink(tmp.name)


# =========================================================================
# Tests for grade_results convenience wrapper
# =========================================================================

class TestGradeResults(unittest.TestCase):
    """Verify the grade_results convenience wrapper orchestrates correctly."""

    def test_grade_results_workflow(self):
        from runner.evaluate import grade_results
        gold_data = [
            {"task_id": "t1", "category": "factual",
             "prompt": "Capital?", "gold": {"answer": "Paris"}},
            {"task_id": "t2", "category": "math",
             "prompt": "2+2?", "gold": {"answer": 4}},
        ]
        preds_data = [
            {"task_id": "t1", "answer": "Paris", "timing_ms": 10.0},
            {"task_id": "t2", "answer": "4", "timing_ms": 5.0},
        ]
        gold_path = _make_temp_json(gold_data)
        preds_path = _make_temp_json(preds_data)
        xlsx_path = gold_path + ".xlsx"  # unique path alongside one temp file
        try:
            report = grade_results(preds_path, gold_path, xlsx_path, verbose=False)
            assert report["overall"]["total"] == 2
            assert report["overall"]["correct"] == 2
            assert report["overall"]["gate_pass"] is True
            assert len(report["per_task"]) == 2
            # XLSX should exist
            assert os.path.exists(xlsx_path)
            assert os.path.getsize(xlsx_path) > 0
        finally:
            os.unlink(gold_path)
            os.unlink(preds_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)


if __name__ == "__main__":
    unittest.main()
