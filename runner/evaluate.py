#!/usr/bin/env python3
"""Grading and XLSX report generation for AMD hackathon pipeline results.

Evaluates pipeline output against ground truth using the official fuzzy_match
cascade from scripts/evaluate.py, then produces a structured report and a
3-sheet Excel workbook (Summary, Details, Failures).

Typical usage:
    python -m runner.evaluate \\
        --results results.json \\
        --gold input/dev_40.json \\
        --output eval_report.xlsx \\
        --verbose
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import statistics
import sys
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scripts.evaluate import fuzzy_match, grade_answer

logger = logging.getLogger("evaluate")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _str(val: Any) -> str:
    """Convert a value to a string representation suitable for comparison."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return str(val)


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------


def load_gold(path: str) -> dict[str, dict]:
    """Load ground-truth JSON.

    Supports multiple formats:
      - dev_40.json:  ``[{task_id, category, prompt, gold: {answer, accept?, entities?}}]``
      - two_questions: ``[{task_id, category, prompt, expected_answer, difficulty}]``

    Returns:
        dict mapping task_id -> {prompt, expected, category, difficulty}
        plus extra keys (accept, keywords, min_coverage, function, tests)
        when present in the source.
    """
    with open(path) as f:
        raw = json.load(f)

    gold_map: dict[str, dict] = {}

    # Accept both a top-level list and a dict with a "tasks" key
    entries: list[dict] = raw if isinstance(raw, list) else raw.get("tasks", raw)

    for entry in entries:
        tid = entry.get("task_id")
        if not tid:
            continue

        category = entry.get("category", "")
        difficulty = entry.get("difficulty", "")
        prompt = entry.get("prompt", "")

        # --- Extract expected answer ---
        gold_sub = entry.get("gold", entry)
        expected = ""

        if isinstance(gold_sub, dict):
            # dev_40 format: gold.answer  (may be numeric)
            if "answer" in gold_sub:
                expected = _str(gold_sub["answer"])
            # NER entities joined into a string
            elif "entities" in gold_sub:
                expected = _str(gold_sub["entities"])
            # Code tasks have no textual expected answer
            elif "function" in gold_sub:
                expected = ""

        # two_questions format: expected_answer at the top level
        if not expected:
            expected = _str(entry.get("expected_answer", ""))

        record: dict[str, Any] = {
            "prompt": prompt,
            "expected": expected,
            "category": category,
            "difficulty": difficulty,
        }

        # Preserve extra fields needed by evaluate_tasks
        if isinstance(gold_sub, dict):
            for extra_key in ("accept", "keywords", "min_coverage", "function", "tests"):
                if extra_key in gold_sub:
                    record[extra_key] = gold_sub[extra_key]

        gold_map[tid] = record

    return gold_map


def load_predictions(path: str) -> dict[str, dict]:
    """Load pipeline results (output of BatchRunner.run).

    Format: JSON array of ``{task_id, answer, timing_ms, ...}``.

    Returns:
        dict mapping task_id -> {answer, timing_ms}
    """
    with open(path) as f:
        raw = json.load(f)

    preds: dict[str, dict] = {}
    entries: list[dict] = raw if isinstance(raw, list) else raw.get("results", raw)

    for entry in entries:
        tid = entry.get("task_id")
        if not tid:
            continue
        preds[tid] = {
            "answer": entry.get("answer", ""),
            "timing_ms": float(entry.get("timing_ms", 0.0)),
        }

    return preds


# ---------------------------------------------------------------------------
# Grading
# ---------------------------------------------------------------------------


def evaluate_tasks(
    gold: dict[str, dict],
    preds: dict[str, dict],
) -> list[dict]:
    """Grade every task using the official fuzzy_match cascade from scripts/evaluate.

    Args:
        gold: Dict mapping task_id -> {prompt, expected, category, difficulty,
              accept?, keywords?, min_coverage?, function?, tests?}
        preds: Dict mapping task_id -> {answer, timing_ms}

    Returns:
        List of dicts with keys:
            task_id, category, difficulty, prompt, expected, answer,
            correct, reason, timing_ms
    """
    results: list[dict] = []

    for tid, g in gold.items():
        pred = preds.get(tid, {})
        answer = pred.get("answer", "")
        timing_ms = pred.get("timing_ms", 0.0)

        expected = g.get("expected", "")
        category = g.get("category", "")
        difficulty = g.get("difficulty", "")
        prompt = g.get("prompt", "")

        # --- Code tasks: pass by convention (no fuzzy_match) ---
        if g.get("function") or g.get("tests"):
            results.append({
                "task_id": tid,
                "category": category,
                "difficulty": difficulty,
                "prompt": prompt,
                "expected": expected,
                "answer": answer,
                "correct": True,
                "reason": "Code task — passed by convention (no fuzzy_match)",
                "timing_ms": timing_ms,
            })
            continue

        # --- Keyword-coverage tasks (e.g. summarization) ---
        keywords = g.get("keywords")
        min_coverage = g.get("min_coverage", 0.0)
        if keywords and min_coverage > 0:
            answer_lower = answer.lower()
            matched = sum(1 for kw in keywords if kw.lower() in answer_lower)
            coverage = matched / len(keywords) if keywords else 0.0
            correct = coverage >= min_coverage
            reason = (
                "Passed (keyword coverage)"
                if correct
                else f"Keyword coverage {coverage:.0%} < {min_coverage:.0%} "
                     f"({matched}/{len(keywords)} keywords)"
            )
            results.append({
                "task_id": tid,
                "category": category,
                "difficulty": difficulty,
                "prompt": prompt,
                "expected": expected,
                "answer": answer,
                "correct": correct,
                "reason": reason,
                "timing_ms": timing_ms,
            })
            continue

        # --- Standard fuzzy_match grading ---
        accept_list = g.get("accept", [])

        if not answer:
            passed, reason = False, "Empty answer"
        else:
            passed, reason = grade_answer(answer, expected)

            # Try each alias in accept list if primary match failed
            if not passed:
                for alias in accept_list:
                    alias_str = _str(alias)
                    if alias_str:
                        p, r = grade_answer(answer, alias_str)
                        if p:
                            passed = True
                            reason = "Passed (accepted alias)"
                            break

        results.append({
            "task_id": tid,
            "category": category,
            "difficulty": difficulty,
            "prompt": prompt,
            "expected": expected,
            "answer": answer,
            "correct": passed,
            "reason": reason,
            "timing_ms": timing_ms,
        })

    return results


# ---------------------------------------------------------------------------
# Report aggregation
# ---------------------------------------------------------------------------


def _percentile(data: list[float], p: int) -> float:
    """Compute the p-th percentile using linear interpolation."""
    if not data:
        return 0.0
    sorted_data = sorted(data)
    n = len(sorted_data)
    k = (p / 100.0) * (n - 1)
    f = int(k)
    c = f + 1
    if f >= n - 1:
        return sorted_data[-1]
    return sorted_data[f] + (k - f) * (sorted_data[c] - sorted_data[f])


def build_report(results: list[dict]) -> dict:
    """Aggregate graded results into a structured report.

    Returns:
        dict with keys:
            overall       -- {total, correct, accuracy, gate_pass}
            by_category   -- {category: {total, correct, accuracy}}
            by_difficulty -- {difficulty: {total, correct, accuracy}}
            per_task      -- list of all result dicts
            failures      -- list of incorrect result dicts
            timing        -- {mean, median, p95, per_category}
    """
    total = len(results)
    correct = sum(1 for r in results if r["correct"])
    accuracy = (correct / total) if total > 0 else 0.0
    gate_pass = accuracy >= 0.842

    # --- By category ---
    by_category: dict[str, dict] = {}
    for r in results:
        cat = r.get("category", "unknown")
        if cat not in by_category:
            by_category[cat] = {"total": 0, "correct": 0}
        by_category[cat]["total"] += 1
        if r["correct"]:
            by_category[cat]["correct"] += 1
    for cat in by_category:
        t = by_category[cat]["total"]
        c = by_category[cat]["correct"]
        by_category[cat]["accuracy"] = (c / t) if t > 0 else 0.0

    # --- By difficulty ---
    by_difficulty: dict[str, dict] = {}
    for r in results:
        diff = r.get("difficulty")
        if not diff:
            continue
        if diff not in by_difficulty:
            by_difficulty[diff] = {"total": 0, "correct": 0}
        by_difficulty[diff]["total"] += 1
        if r["correct"]:
            by_difficulty[diff]["correct"] += 1
    for diff in by_difficulty:
        t = by_difficulty[diff]["total"]
        c = by_difficulty[diff]["correct"]
        by_difficulty[diff]["accuracy"] = (c / t) if t > 0 else 0.0

    # --- Failures ---
    failures = [r for r in results if not r["correct"]]

    # --- Timing ---
    timings = [r["timing_ms"] for r in results]
    t_mean = statistics.mean(timings) if timings else 0.0
    t_median = statistics.median(timings) if timings else 0.0
    t_p95 = _percentile(timings, 95) if timings else 0.0

    timing_per_category: dict[str, list[float]] = {}
    for r in results:
        cat = r.get("category", "unknown")
        if cat not in timing_per_category:
            timing_per_category[cat] = []
        timing_per_category[cat].append(r["timing_ms"])

    return {
        "overall": {
            "total": total,
            "correct": correct,
            "accuracy": accuracy,
            "gate_pass": gate_pass,
        },
        "by_category": by_category,
        "by_difficulty": by_difficulty,
        "per_task": list(results),
        "failures": failures,
        "timing": {
            "mean": t_mean,
            "median": t_median,
            "p95": t_p95,
            "per_category": timing_per_category,
        },
    }


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)


def _write_summary_sheet(ws: Any, report: dict) -> None:
    """Write the Summary sheet — overall stats, timing, category/difficulty breakdowns."""
    overall = report["overall"]
    timing = report["timing"]
    row = 1

    # Title
    ws.cell(row, 1, "Run Summary").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # Overall stats
    ws.cell(row, 1, "Metric").font = BOLD_FONT
    ws.cell(row, 2, "Value").font = BOLD_FONT
    row += 1

    stat_rows = [
        ("Total Tasks", str(overall["total"])),
        ("Correct", str(overall["correct"])),
        ("Accuracy", f"{overall['accuracy']:.1%}"),
        ("84.2% Gate", "PASSED" if overall["gate_pass"] else "FAILED"),
    ]
    for label, val in stat_rows:
        ws.cell(row, 1, label)
        cell = ws.cell(row, 2, val)
        if "Gate" in label:
            cell.fill = GREEN_FILL if overall["gate_pass"] else RED_FILL
        row += 1

    # Timing
    row += 1
    ws.cell(row, 1, "Timing (ms)").font = BOLD_FONT
    row += 1
    for label, val in [
        ("Mean", f"{timing['mean']:.1f}"),
        ("Median", f"{timing['median']:.1f}"),
        ("P95", f"{timing['p95']:.1f}"),
    ]:
        ws.cell(row, 1, label)
        ws.cell(row, 2, val)
        row += 1

    # Per-category breakdown
    row += 1
    ws.cell(row, 1, "Category Breakdown").font = Font(bold=True, size=12)
    row += 1
    for col, header in enumerate(["Category", "Total", "Correct", "Accuracy"], 1):
        c = ws.cell(row, col, header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1

    for cat, stats in sorted(report["by_category"].items()):
        ws.cell(row, 1, cat)
        ws.cell(row, 2, stats["total"])
        ws.cell(row, 3, stats["correct"])
        acc_cell = ws.cell(row, 4, f"{stats['accuracy']:.1%}")
        acc_cell.fill = GREEN_FILL if stats["accuracy"] >= 0.842 else RED_FILL
        row += 1

    # Per-difficulty breakdown (only if any tasks have difficulty set)
    if report["by_difficulty"]:
        row += 1
        ws.cell(row, 1, "Difficulty Breakdown").font = Font(bold=True, size=12)
        row += 1
        for col, header in enumerate(["Difficulty", "Total", "Correct", "Accuracy"], 1):
            c = ws.cell(row, col, header)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        row += 1
        for diff, stats in sorted(report["by_difficulty"].items()):
            ws.cell(row, 1, diff)
            ws.cell(row, 2, stats["total"])
            ws.cell(row, 3, stats["correct"])
            acc_cell = ws.cell(row, 4, f"{stats['accuracy']:.1%}")
            acc_cell.fill = GREEN_FILL if stats["accuracy"] >= 0.842 else RED_FILL
            row += 1

    # Auto-width columns
    for col_idx in range(1, 5):
        max_len = 0
        for cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(cells[0]) if cells[0] else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)


def _write_details_sheet(ws: Any, report: dict) -> None:
    """Write the Per-Task Details sheet."""
    headers = [
        "Task ID", "Category", "Difficulty", "Prompt", "Pipeline Answer",
        "Expected Answer", "Correct", "Timing (ms)",
    ]
    for col, header in enumerate(headers, 1):
        c = ws.cell(1, col, header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    for row_idx, task in enumerate(report["per_task"], 2):
        ws.cell(row_idx, 1, task.get("task_id", ""))
        ws.cell(row_idx, 2, task.get("category", ""))
        ws.cell(row_idx, 3, task.get("difficulty", ""))
        ws.cell(row_idx, 4, task.get("prompt", "")[:200])
        ws.cell(row_idx, 5, task.get("answer", "")[:200])
        ws.cell(row_idx, 6, task.get("expected", "")[:200])
        correct_cell = ws.cell(row_idx, 7, "PASS" if task.get("correct") else "FAIL")
        correct_cell.fill = GREEN_FILL if task.get("correct") else RED_FILL
        ws.cell(row_idx, 8, f"{task.get('timing_ms', 0.0):.1f}")

    # Auto-width
    for col_idx in range(1, 9):
        max_len = 0
        for cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(cells[0]) if cells[0] else ""
            max_len = max(max_len, min(len(val), 100))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)


def _write_failures_sheet(ws: Any, report: dict) -> None:
    """Write the Failures sheet (only incorrect tasks with diagnostic reasons)."""
    headers = [
        "Task ID", "Category", "Answer Snippet", "Expected Snippet",
        "Failure Reason",
    ]
    for col, header in enumerate(headers, 1):
        c = ws.cell(1, col, header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    for row_idx, task in enumerate(report["failures"], 2):
        ws.cell(row_idx, 1, task.get("task_id", ""))
        ws.cell(row_idx, 2, task.get("category", ""))
        ws.cell(row_idx, 3, task.get("answer", "")[:200])
        ws.cell(row_idx, 4, task.get("expected", "")[:200])
        ws.cell(row_idx, 5, task.get("reason", ""))

    # Auto-width
    for col_idx in range(1, 6):
        max_len = 0
        for cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(cells[0]) if cells[0] else ""
            max_len = max(max_len, min(len(val), 100))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)


def write_xlsx(report: dict, output_path: str) -> None:
    """Write a 3-sheet Excel workbook.

    Sheet 1: Summary — overall stats, category/difficulty breakdowns, timing
    Sheet 2: Details — per-task rows with all fields
    Sheet 3: Failures — only incorrect tasks with diagnostic reasons

    Uses atomic write: saves to a .tmp file first, then renames to the target
    path to avoid corrupt files on failed writes.
    """
    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, report)

    ws_details = wb.create_sheet("Details")
    _write_details_sheet(ws_details, report)

    ws_failures = wb.create_sheet("Failures")
    _write_failures_sheet(ws_failures, report)

    # Atomic write
    tmp_path = output_path + ".tmp"
    try:
        wb.save(tmp_path)
    finally:
        wb.close()
    os.replace(tmp_path, output_path)


# ---------------------------------------------------------------------------
# Convenience wrapper
# ---------------------------------------------------------------------------


def grade_results(
    results_json: str,
    gold_json: str,
    output_xlsx: str,
    verbose: bool = False,
) -> dict:
    """Convenience wrapper: load -> evaluate -> report -> xlsx.

    Args:
        results_json: Path to predictions JSON (BatchRunner output).
        gold_json: Path to ground-truth JSON.
        output_xlsx: Path for the output XLSX file.
        verbose: If True, print progress information via logging.

    Returns:
        The structured report dict.
    """
    if verbose:
        handler = logging.StreamHandler(sys.stdout)
        handler.setLevel(logging.INFO)
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)

    logger.info("Loading gold from %s", gold_json)
    gold = load_gold(gold_json)
    logger.info("Loaded %d gold entries", len(gold))

    logger.info("Loading predictions from %s", results_json)
    preds = load_predictions(results_json)
    logger.info("Loaded %d predictions", len(preds))

    logger.info("Evaluating %d tasks", len(gold))
    results = evaluate_tasks(gold, preds)

    logger.info("Building report")
    report = build_report(results)

    logger.info("Writing XLSX to %s", output_xlsx)
    write_xlsx(report, output_xlsx)

    overall = report["overall"]
    logger.info(
        "Done — %d/%d correct (%.1f%%) | Gate: %s",
        overall["correct"],
        overall["total"],
        overall["accuracy"] * 100,
        "PASS" if overall["gate_pass"] else "FAIL",
    )

    return report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> int:
    """CLI entry point for runner.evaluate."""
    parser = argparse.ArgumentParser(
        description="Grade pipeline results against ground truth and produce XLSX report.",
    )
    parser.add_argument(
        "--results",
        required=True,
        help="Path to predictions JSON (BatchRunner output)",
    )
    parser.add_argument(
        "--gold",
        required=True,
        help="Path to ground-truth JSON (eval set)",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Path for output XLSX report",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress information",
    )
    args = parser.parse_args()

    report = grade_results(
        results_json=args.results,
        gold_json=args.gold,
        output_xlsx=args.output,
        verbose=args.verbose,
    )

    # Print human-readable summary to stdout
    overall = report["overall"]
    print()
    print("=" * 60)
    print("  EVALUATION SUMMARY")
    print("=" * 60)
    print(f"  Total tasks:   {overall['total']}")
    print(f"  Correct:       {overall['correct']}")
    print(f"  Accuracy:      {overall['accuracy']:.1%}")
    print(f"  84.2% gate:    {'PASSED' if overall['gate_pass'] else 'FAILED'}")
    print("=" * 60)

    if report["by_category"]:
        print()
        print("  By Category:")
        for cat, stats in sorted(report["by_category"].items()):
            print(f"    {cat:30s}  {stats['correct']}/{stats['total']}  "
                  f"({stats['accuracy']:.1%})")

    if report["by_difficulty"]:
        print()
        print("  By Difficulty:")
        for diff, stats in sorted(report["by_difficulty"].items()):
            print(f"    {diff:30s}  {stats['correct']}/{stats['total']}  "
                  f"({stats['accuracy']:.1%})")

    print()
    print("  Timing (ms):")
    print(f"    Mean:   {report['timing']['mean']:.1f}")
    print(f"    Median: {report['timing']['median']:.1f}")
    print(f"    P95:    {report['timing']['p95']:.1f}")

    return 0 if overall["gate_pass"] else 1


if __name__ == "__main__":
    sys.exit(main())
