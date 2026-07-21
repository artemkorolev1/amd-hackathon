#!/usr/bin/env python3
"""
Standalone pipeline for Google Colab (T4 GPU).
Packages the AMD ACT II hackathon pipeline into a single self-contained script.

Usage:
    python3 colab_pipeline.py --eval data/eval/training-v1.json --model /path/to/model.gguf
    python3 colab_pipeline.py --eval data/eval/validation-v1.json --model /path/to/model.gguf --gpu-layers -1

Outputs:
    results/colab_run_TIMESTAMP.json       Full per-question results
    results/colab_run_TIMESTAMP_summary.txt Human-readable summary
"""

import argparse
import csv
import hashlib
import json
import logging
import os
import re
import sys
import threading
import time
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    stream=sys.stderr,
)
logger = logging.getLogger("colab_pipeline")

# =========================================================================
# CONFIG
# =========================================================================

DEFAULT_MODEL = "/models/qwen2.5-1.5b-instruct-q4_k_m.gguf"
N_GPU_LAYERS = -1
N_CTX = 2048
N_THREADS = 4
MAX_TOKENS = 256
COMPLEXITY_SIMPLE_MAX = 0.3
CONSENSUS_SAMPLES = 1

DETERMINISTIC_CATEGORIES = {"math", "logic", "sentiment", "ner", "factual", "code_debug"}
NAKED_CATEGORIES = {"ner", "summarization", "factual", "logic", "math"}

CATEGORIES_8WAY = [
    "code_debug", "code_gen", "factual", "logic",
    "math", "ner", "sentiment", "summarization",
]

PRIORITY = {
    "code_debug": 8, "code_gen": 7, "math": 6, "logic": 5,
    "sentiment": 4, "ner": 3, "summarization": 2, "factual": 1,
}

DET_CATEGORY_MAP = {
    "math": "math_arithmetic",
    "logic": "logical_reasoning",
    "sentiment": "sentiment",
    "ner": "named_entity_recognition",
    "factual": "other_complex",
    "code_debug": "code_debugging",
    "code_gen": "code_debugging",
    "summarization": "summarization",
    "general": "other_complex",
}

HUMAN_NAMES = {
    "code_debug": "Code Debugging", "code_gen": "Code Generation",
    "factual": "Factual Knowledge", "logic": "Logical Reasoning",
    "math": "Math Reasoning", "ner": "Named Entity Recognition",
    "sentiment": "Sentiment", "summarization": "Summarisation",
}


# =========================================================================
# DETERMINISTIC SOLVERS
# =========================================================================

def solve_arithmetic(prompt: str, category: str = "") -> Optional[str]:
    """Solve arithmetic expressions using safe eval."""
    prompt_clean = prompt.strip()
    # Pure arithmetic: numbers and operators only
    arith_pattern = re.compile(
        r'^[\d\s\+\-\*\/\(\)\%\.\,]+$'
    )
    # Check if it's a calculation request
    calc_keywords = re.compile(
        r'(calculate|compute|what\s+is|result\s+of|solve|evaluate|simplify)\b',
        re.IGNORECASE
    )
    if not calc_keywords.search(prompt_clean) and not arith_pattern.match(prompt_clean):
        return None

    # Extract the expression
    nums = re.findall(r'[\d]+(?:\s*[\+\-\*\/\(\)]\s*[\d]+)+', prompt_clean)
    if not nums:
        nums = re.findall(r'[\d]+(?:\s*[\+\-\*\/\(\)]\s*[\d]+)*\s*=', prompt_clean)
    if not nums:
        nums = re.findall(r'[\d\s\+\-\*\/\(\)\%]+', prompt_clean)
    if not nums:
        return None

    expr = nums[0].strip().rstrip('=')
    try:
        # Safe eval - only allow basic math
        result = eval(expr, {"__builtins__": {}}, {})
        if isinstance(result, (int, float)):
            return f"{result:.4f}".rstrip('0').rstrip('.') if isinstance(result, float) else str(result)
    except Exception:
        pass
    return None


def solve_logic(prompt: str, category: str = "") -> Optional[str]:
    """Solve simple logic puzzles (truth-tellers, liars, syllogisms)."""
    # Truth-teller / liar patterns
    truth_teller = re.compile(
        r'(truth.?teller|liar|always\s+tells?\s+(the\s+)?truth|always\s+lies)',
        re.IGNORECASE
    )
    if not truth_teller.search(prompt):
        return None
    return None  # Too complex for deterministic - let LLM handle


def solve_sentiment(prompt: str, category: str = "") -> Optional[str]:
    """Simple sentiment detection for clear cases."""
    prompt_lower = prompt.lower()
    sentiment_words = {
        'positive': {'happy', 'great', 'excellent', 'wonderful', 'amazing', 'fantastic',
                     'love', 'beautiful', 'perfect', 'brilliant', 'delighted', 'pleased'},
        'negative': {'terrible', 'awful', 'horrible', 'hate', 'bad', 'worst',
                     'disappointed', 'sad', 'angry', 'frustrating', 'poor', 'dreadful'},
        'neutral': {'the', 'this', 'that', 'it', 'is', 'was', 'are', 'were', 'has', 'have',
                    'said', 'says', 'according', 'reported'}
    }
    # Count sentiment signals
    pos = sum(1 for w in sentiment_words['positive'] if w in prompt_lower)
    neg = sum(1 for w in sentiment_words['negative'] if w in prompt_lower)

    total = pos + neg
    if total >= 3:
        ratio = pos / total
        if ratio > 0.66:
            return "positive"
        elif ratio < 0.33:
            return "negative"
    return None


def solve_ner(prompt: str, category: str = "") -> Optional[str]:
    """Simple NER for common patterns."""
    entities = []
    # Person names (Captain Formals pattern)
    persons = re.findall(r'(?:Dr\.|Mr\.|Mrs\.|Ms\.|Prof\.)\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*', prompt)
    entities.extend(persons)

    # Organizations
    orgs = re.findall(r'\b(?:[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\s+(?:Inc|Corp|LLC|Ltd|Group|Technologies|Systems|Pharmaceuticals|Laboratories|University|Institute|Hospital|Airlines))\b', prompt)
    entities.extend(orgs)

    # Dates
    dates = re.findall(r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4}\b', prompt)
    entities.extend(dates)

    if entities:
        return ", ".join(entities)
    return None


def solve_factual_qa(prompt: str, category: str = "") -> Optional[str]:
    """Simple factual QA for common knowledge."""
    prompt_lower = prompt.lower().strip()

    # Capital city questions
    capital_q = re.compile(r'(?:what\s+is\s+the\s+capital|capital\s+of)\s+(\w+(?:\s+\w+)?)\s*\?', re.IGNORECASE)
    m = capital_q.search(prompt)
    if m:
        country = m.group(1).strip().lower()
        capitals = {
            "france": "Paris", "germany": "Berlin", "italy": "Rome",
            "spain": "Madrid", "uk": "London", "united kingdom": "London",
            "japan": "Tokyo", "china": "Beijing", "india": "New Delhi",
            "australia": "Canberra", "canada": "Ottawa", "brazil": "Brasília",
            "russia": "Moscow", "usa": "Washington, D.C.", "united states": "Washington, D.C.",
            "united states of america": "Washington, D.C.", "egypt": "Cairo",
            "nigeria": "Abuja", "kenya": "Nairobi", "south africa": "Pretoria",
        }
        if country in capitals:
            return capitals[country]

    return None


def solve_code_debugging(prompt: str, category: str = "") -> Optional[str]:
    """Fix common bug patterns in code."""
    return None  # Too complex to be reliable

# Map deterministic solver names to functions
DET_SOLVER_FNS = {
    "math_arithmetic": solve_arithmetic,
    "logical_reasoning": solve_logic,
    "sentiment": solve_sentiment,
    "named_entity_recognition": solve_ner,
    "other_complex": solve_factual_qa,
    "code_debugging": solve_code_debugging,
}


# =========================================================================
# 8-WAY CATEGORY CLASSIFIER (regex-based)
# =========================================================================

def _score_code_debug(prompt: str) -> float:
    """Score for code_debugging category."""
    score = 0.0
    if re.search(r'\b(bug|error|fix|debug|issue|crash|broken|incorrect|wrong|fail|not work|debugging)\b',
                 prompt, re.IGNORECASE):
        score += 3.0
    if re.search(r'\b(exception|traceback|typeerror|valueerror|keyerror|indexerror|attributeerror|importerror)\b',
                 prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(def\s+\w+\s*\(|class\s+\w+|import\s+\w+|return\s+|print\s*\()', prompt):
        score += 2.0
    if re.search(r'\b(test|output|result|expected|actual)\b.*\b(different|mismatch|wrong|incorrect)\b',
                 prompt, re.IGNORECASE):
        score += 1.5
    if re.search(r'\b(bug\s+in\s+the\s+(Python|code|function)|the\s+bug\sin)\b', prompt, re.IGNORECASE):
        score += 1.0
    # Large code block
    code_lines = len(re.findall(r'^\s*(def |class |import |from |return |if |for |while |with |try:)', prompt, re.MULTILINE))
    score += min(code_lines * 0.5, 2.0)
    return score


def _score_code_gen(prompt: str) -> float:
    """Score for code_generation category."""
    score = 0.0
    if re.search(r'\b(write|create|implement|generate|build|code|program|function|method|class)\b',
                 prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(Write\s+a\s+(Python|Java|JavaScript|C\+\+|function)|Implement\s+(a\s+)?(function|class|method|algorithm))\b',
                 prompt):
        score += 2.0
    if re.search(r'\b(return|print|input|def |class |import )', prompt):
        score += 1.5
    if re.search(r'\b(Write\s+a\s+function|write\s+the\s+code|generate\s+code|implement\s+(a|the)\s+function)\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    if re.search(r'\b(using\s+(Python|recursion|iteration|list\s+comprehension|dictionary|set))\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    return score


def _score_factual(prompt: str) -> float:
    """Score for factual_knowledge category."""
    score = 0.0
    if re.search(r'\b(what|who|when|where|why|which|how)\b.*\?', prompt):
        score += 2.0
    if re.search(r'\b(known\s+as|called|also\s+known\s+as|referred\s+to\s+as)\b', prompt, re.IGNORECASE):
        score += 1.5
    if re.search(r'\b(fact|information|knowledge|defined\s+as|meaning|definition|explain|describe)\b',
                 prompt, re.IGNORECASE):
        score += 1.5
    if re.search(r'\b(country|capital|city|river|mountain|ocean|planet|star|element|chemical)\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    if re.search(r'\b(history|discovered|invented|founded|born|died|president|king|queen|leader)\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    if re.search(r'\b(What\s+is|Who\s+(is|was)|When\s+(did|was)|Where\s+(is|was)|How\s+(many|much|far|long|old))\b',
                 prompt):
        score += 1.0
    return score


def _score_logic(prompt: str) -> float:
    """Score for logical_reasoning category."""
    score = 0.0
    if re.search(r'\b(if\s+.*\bthen|either\s+.*\bor|neither\s+.*\bnor|all\s+.*\bare|no\s+.*\bare)\b',
                 prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(logic|reasoning|conclusion|infer|deduce|premise|assumption|implication)\b',
                 prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(must\s+be|cannot\s+be|could\s+be|always|never|sometimes|necessarily|possibly)\b',
                 prompt, re.IGNORECASE):
        score += 1.5
    if re.search(r'\b(truth.?teller|liar|knights?|knaves?|puzzle|riddle)\b', prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(statement|contradiction|paradox|syllogism|valid|invalid)\b', prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(which\s+of\s+the\s+following|choose|select|pick)\b', prompt, re.IGNORECASE):
        score += 1.0
    if re.search(r'\b(only|every|some|none|all|no\s+one|everyone)\b.*\b(is|are|can|cannot|will|must)\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    return score


def _score_math(prompt: str) -> float:
    """Score for math_reasoning category."""
    score = 0.0
    # Numbers and math operators
    if re.search(r'[\d]+[\s]*[\+\-\*\/\^\%][\s]*[\d]+', prompt):
        score += 2.0
    if re.search(r'\b(calculate|compute|solve|evaluate|simplify|find|determine)\b.*[\d]', prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(divided\s+by|multiplied\s+by|plus|minus|times|over|percent|percentage|fraction|decimal)\b',
                 prompt, re.IGNORECASE):
        score += 1.5
    if re.search(r'\b(distance|speed|rate|time|volume|area|length|width|height|radius|diameter)\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    if re.search(r'\b(equation|formula|theorem|proof|derivative|integral|sum|product|function)\b',
                 prompt, re.IGNORECASE):
        score += 1.5
    if re.search(r'\b(probability|permutation|combination|mean|median|mode|standard\s+deviation|variance)\b',
                 prompt, re.IGNORECASE):
        score += 1.5
    if re.search(r'\b[A-Za-z]\s*=\s*[\d]', prompt):
        score += 1.0
    if re.search(r'\b(what\s+is\s+the\s+(value|result|sum|product|difference|average|total))\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    return score


def _score_ner(prompt: str) -> float:
    """Score for named_entity_recognition category."""
    score = 0.0
    # Named entities (Capitalized words)
    capitalized = len(re.findall(r'\b[A-Z][a-z]+\b', prompt))
    if capitalized > 5:
        score += 2.0
    if re.search(r'\b(identify|find|extract|list|locate|tag|label|recognize)\b.*\b(entity|name|person|organization|location|date|place)\b',
                 prompt, re.IGNORECASE):
        score += 2.5
    if re.search(r'\b(NER|named\s+entity|entity\s+recognition|entity\s+extraction|entity\s+tagging)\b',
                 prompt, re.IGNORECASE):
        score += 3.0
    if re.search(r'\b(Dr\.|Mr\.|Mrs\.|Ms\.|Prof\.|CEO|President|Director|Professor|Doctor)\b', prompt):
        score += 1.5
    if re.search(r'\b(Inc|Corp|LLC|Ltd|University|Hospital|Foundation|Organization|Agency|Department)\b',
                 prompt):
        score += 1.0
    if re.search(r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2}',
                 prompt):
        score += 1.0
    if re.search(r'\b(patient|diagnosis|symptom|treatment|disease|syndrome|condition)\b', prompt, re.IGNORECASE):
        score += 1.0
    return score


def _score_sentiment(prompt: str) -> float:
    """Score for sentiment_classification category."""
    score = 0.0
    positive_words = re.findall(r'\b(good|great|excellent|amazing|fantastic|wonderful|happy|love|beautiful|perfect|brilliant|delighted|pleased|awesome|positive)\b', prompt, re.IGNORECASE)
    negative_words = re.findall(r'\b(bad|terrible|awful|horrible|hate|worst|disappointed|sad|angry|frustrating|poor|dreadful|negative|unpleasant|disgusting)\b', prompt, re.IGNORECASE)
    pos_count = len(positive_words)
    neg_count = len(negative_words)

    if re.search(r'\b(sentiment|emotion|feeling|opinion|mood|reaction|attitude)\b', prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(classify|determine|identify|analyze|detect|assess|evaluate)\b.*\b(sentiment|emotion|feeling|opinion)\b',
                 prompt, re.IGNORECASE):
        score += 2.0
    if re.search(r'\b(positive|negative|neutral)\b', prompt, re.IGNORECASE):
        score += 1.5
    score += min(pos_count * 0.5, 1.5)
    score += min(neg_count * 0.5, 1.5)
    if re.search(r'\b(movie|review|product|service|customer|experience|feedback|comment)\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    return score


def _score_summarization(prompt: str) -> float:
    """Score for summarisation category."""
    score = 0.0
    if re.search(r'\b(summarize|summarise|summary|brief|condense|shorten|synopsis|abstract|overview|gist|tl;dr|tl dr)\b',
                 prompt, re.IGNORECASE):
        score += 3.0
    if len(prompt.split()) > 100:
        score += 2.0
    if re.search(r'\b(in\s+one\s+sentence|in\s+a\s+few\s+sentences|in\s+your\s+own\s+words|in\s+short|concisely|briefly)\b',
                 prompt, re.IGNORECASE):
        score += 1.5
    if re.search(r'\b(article|text|passage|paragraph|document|report|paper|chapter|section|email|letter|story)\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    if re.search(r'\b(extract|key\s+points|main\s+idea|main\s+points|essential|core)\b',
                 prompt, re.IGNORECASE):
        score += 1.0
    # Long text indicates need for summarization
    word_count = len(prompt.split())
    if word_count > 150:
        score += 1.0
    return score


SCORERS: Dict[str, Any] = {
    "code_debug": _score_code_debug,
    "code_gen": _score_code_gen,
    "factual": _score_factual,
    "logic": _score_logic,
    "math": _score_math,
    "ner": _score_ner,
    "sentiment": _score_sentiment,
    "summarization": _score_summarization,
}


def classify(prompt: str) -> Tuple[str, float, Dict[str, float]]:
    """8-way category classifier. Returns (category, confidence, all_scores)."""
    scores: Dict[str, float] = {}
    for cat, scorer in SCORERS.items():
        scores[cat] = scorer(prompt)
    # Highest score wins, ties broken by priority
    best_cat = max(scores, key=lambda c: (scores[c], PRIORITY.get(c, 0)))
    second_best = sorted(scores.items(), key=lambda x: (-x[1], -PRIORITY.get(x[0], 0)))[1]
    confidence = scores[best_cat] - second_best[1]
    return best_cat, confidence, scores


# =========================================================================
# COMPLEXITY SCORER
# =========================================================================

def score_complexity(prompt: str, category: str = "factual") -> float:
    """Per-category complexity score (0.0–1.0)."""
    if category == "code_gen":
        return _complexity_code_gen(prompt)
    elif category == "code_debug":
        return _complexity_code_debug(prompt)
    elif category == "math":
        return _complexity_math(prompt)
    elif category == "logic":
        return _complexity_logic(prompt)
    elif category == "factual":
        return _complexity_factual(prompt)
    elif category == "sentiment":
        return _complexity_sentiment(prompt)
    elif category == "ner":
        return _complexity_ner(prompt)
    elif category == "summarization":
        return _complexity_summarization(prompt)
    return 0.5


def _complexity_code_gen(prompt: str) -> float:
    signals = 0.0
    func_count = len(re.findall(r'\b(def\s+\w+\s*\(|function\s+\w+|class\s+\w+)', prompt))
    loop_depth = len(re.findall(r'\b(for|while)\s+\w', prompt))
    import_count = len(re.findall(r'\b(import|from)\s+\w+', prompt))
    code_len = len(re.findall(r'^.*$', prompt, re.MULTILINE))
    signals += min(func_count * 0.15, 0.3)
    signals += min(loop_depth * 0.1, 0.2)
    signals += min(import_count * 0.1, 0.2)
    signals += min(code_len / 500, 0.3)
    return min(signals, 1.0)


def _complexity_code_debug(prompt: str) -> float:
    signals = 0.0
    error_count = len(re.findall(r'\b(error|bug|exception|issue|wrong|incorrect|fail)', prompt, re.IGNORECASE))
    code_size = len(re.findall(r'^.*$', prompt, re.MULTILINE))
    signals += min(error_count * 0.15, 0.4)
    signals += min(code_size / 400, 0.4)
    sub_funcs = len(re.findall(r'\b(def\s+\w+\s*\()', prompt))
    signals += min(sub_funcs * 0.1, 0.2)
    return min(signals, 1.0)


def _complexity_math(prompt: str) -> float:
    signals = 0.0
    ops = len(re.findall(r'[\+\-\*\/\^\%]', prompt))
    paren_depth = max(len(re.findall(r'\(', prompt)), len(re.findall(r'\)', prompt)))
    proof = 1 if re.search(r'\b(proof|theorem|lemma|axiom|prove|show\s+that)\b', prompt, re.IGNORECASE) else 0
    vars = len(set(re.findall(r'\b[A-Za-z]\b', prompt)))
    signals += min(ops * 0.05, 0.2)
    signals += min(paren_depth * 0.05, 0.2) if paren_depth > 3 else 0
    signals += proof * 0.3
    signals += min(vars * 0.02, 0.3)
    return min(signals, 1.0)


def _complexity_logic(prompt: str) -> float:
    signals = 0.0
    conditions = len(re.findall(r'\b(if|when|unless|provided|given\s+that)\b', prompt, re.IGNORECASE))
    negations = len(re.findall(r'\b(not|no|none|never|neither|nor|without)\b', prompt, re.IGNORECASE))
    quantifiers = len(re.findall(r'\b(all|every|some|any|none|no\s+one|each|both|either|neither|most)\b', prompt, re.IGNORECASE))
    signals += min(conditions * 0.1, 0.3)
    signals += min(negations * 0.1, 0.2)
    signals += min(quantifiers * 0.1, 0.2)
    puzzle = 1 if re.search(r'\b(knights?|knaves?|truth.?teller|liar|puzzle)', prompt, re.IGNORECASE) else 0
    signals += puzzle * 0.3
    return min(signals, 1.0)


def _complexity_factual(prompt: str) -> float:
    signals = 0.0
    specificity = len(re.findall(r'\b(January|February|March|April|May|June|July|August|September|October|November|December|\d{4}|19\d{2}|20\d{2})\b', prompt))
    multi_hop = len(re.findall(r'\b(because|therefore|however|although|despite|while|since|as\s+a\s+result)\b', prompt, re.IGNORECASE))
    ambiguity = len(re.findall(r'\b(might|maybe|possibly|probably|likely|unclear|unknown|uncertain|ambiguous|controversial)\b', prompt, re.IGNORECASE))
    signals += min(specificity * 0.1, 0.3)
    signals += min(multi_hop * 0.15, 0.3)
    signals += min(ambiguity * 0.15, 0.2)
    word_count = len(prompt.split())
    signals += min(word_count / 200, 0.2)
    return min(signals, 1.0)


def _complexity_sentiment(prompt: str) -> float:
    signals = 0.0
    text_len = len(prompt.split())
    signals += min(text_len / 200, 0.4)
    subtle = 1 if re.search(r'\b(sarcasm|irony|hedging|subtle|nuance|mixed|conflicting|contradictory)\b', prompt, re.IGNORECASE) else 0
    signals += subtle * 0.3
    entities = len(re.findall(r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\b', prompt))
    if entities > 3:
        signals += 0.2
    explicit_marker = 1 if re.search(r'\b(classify|determine|identify)\b.*\b(sentiment|emotion|feeling)\b', prompt, re.IGNORECASE) else 0
    if explicit_marker:
        signals -= 0.1
    return max(0.0, min(signals, 1.0))


def _complexity_ner(prompt: str) -> float:
    signals = 0.0
    entity_density = len(re.findall(r'\b[A-Z][a-z]+\b', prompt)) / max(len(prompt.split()), 1)
    signals += min(entity_density * 2.0, 0.4)
    text_len = len(prompt.split())
    signals += min(text_len / 200, 0.3)
    domain = 1 if re.search(r'\b(biomedical|financial|legal|clinical|technical|scientific|medical)\b', prompt, re.IGNORECASE) else 0
    signals += domain * 0.3
    return min(signals, 1.0)


def _complexity_summarization(prompt: str) -> float:
    signals = 0.0
    word_count = len(prompt.split())
    signals += min(word_count / 500, 0.4)
    multi_source = len(re.findall(r'\b(however|but|whereas|although|on\s+the\s+(other\s+)?hand|in\s+contrast|conversely)\b', prompt, re.IGNORECASE))
    signals += min(multi_source * 0.1, 0.3)
    bullet = 1 if re.search(r'\b(bullet|point|list|itemize|enumerate|numbered)\b', prompt, re.IGNORECASE) else 0
    signals += bullet * 0.2
    return min(signals, 1.0)


# =========================================================================
# DYNAMIC PROMPTS
# =========================================================================

_ANTI_PREAMBLE = (
    " English only. Start with the answer directly — no greeting, "
    "no 'I will', no 'The user asks', no meta-commentary. "
    "First word = the answer."
)

_CATEGORY_PROMPTS: dict[str, dict[str, str]] = {
    "code_gen": {
        "low": "You are a code generation assistant. Generate clean, correct code. Answer with just the code.",
        "medium": "You are a code generation assistant. Generate working, well-structured code.",
        "high": "You are a code generation assistant. Generate correct, efficient code with proper edge case handling.",
    },
    "code_debug": {
        "low": "You are a code debugging assistant. Identify and fix the bug. Answer with the corrected code.",
        "medium": "You are a code debugging assistant. Identify the bug and provide the corrected version.",
        "high": "You are a code debugging assistant. Find all bugs, explain the fix, provide corrected code.",
    },
    "math": {
        "low": "You are a math assistant. Solve the problem step by step. Answer with the final result.",
        "medium": "You are a math assistant. Solve with clear reasoning. End with the answer.",
        "high": "You are a math assistant. Work through the problem systematically. Final answer on its own line.",
    },
    "logic": {
        "low": "You are a logic reasoning assistant. Determine the correct answer concisely.",
        "medium": "You are a logic reasoning assistant. Think step by step and state your conclusion.",
        "high": "You are a logic reasoning assistant. Reason systematically and provide the final answer.",
    },
    "factual": {
        "low": "You are a factual knowledge assistant. Answer accurately and concisely.",
        "medium": "You are a factual knowledge assistant. Provide accurate, well-sourced information.",
        "high": "You are a factual knowledge assistant. Give a comprehensive, accurate answer.",
    },
    "sentiment": {
        "low": "You are a sentiment classifier. Respond with exactly: positive, negative, or neutral.",
        "medium": "You are a sentiment classifier. Analyze the sentiment and respond with one word: positive, negative, or neutral.",
        "high": "You are a sentiment analyst. Consider tone, context, and nuance. Respond with one label: positive, negative, or neutral.",
    },
    "ner": {
        "low": "You are an NER system. List all named entities found in the text.",
        "medium": "You are an NER system. Extract all named entities: persons, organizations, locations, dates.",
        "high": "You are an NER system. Extract and label all named entities. Format: Type: Entity.",
    },
    "summarization": {
        "low": "You are a summarization assistant. Summarize the text concisely.",
        "medium": "You are a summarization assistant. Provide a clear summary covering key points.",
        "high": "You are a summarization assistant. Provide a comprehensive summary preserving all key information.",
    },
}


def build_system_prompt(category: str, complexity_score: float = 0.5,
                        custom_instructions: str = "") -> str:
    """Build a category-and-complexity-aware system prompt."""
    tier = "low" if complexity_score < 0.3 else ("medium" if complexity_score < 0.6 else "high")
    prompts = _CATEGORY_PROMPTS.get(category, _CATEGORY_PROMPTS.get("factual"))
    base = prompts.get(tier, prompts.get("medium", ""))
    result = base + _ANTI_PREAMBLE
    if custom_instructions:
        result += " " + custom_instructions
    return result


def get_max_tokens(category: str, complexity_score: float = 0.5) -> int:
    """Per-category max generation tokens."""
    tok_map = {
        "code_gen": 250, "code_debug": 200, "math": 200,
        "logic": 200, "factual": 200, "sentiment": 60,
        "ner": 120, "summarization": 120,
    }
    base = tok_map.get(category, 150)
    if complexity_score > 0.6:
        return min(int(base * 1.5), 400)
    return base


NER_ONE_SHOT_EXAMPLE = (
    "Example input: 'Dr. John Smith from Pfizer met with Prof. Jane Doe at Stanford University on March 15, 2024.'\n"
    "Example output: Person: Dr. John Smith, Prof. Jane Doe | Organization: Pfizer, Stanford University | Date: March 15, 2024"
)


def build_merged_prompt(primary_category: str, secondary_category: str,
                        complexity_score: float = 0.5,
                        custom_instructions: str = "") -> str:
    """Build prompt when classifier has low confidence between 2 categories."""
    primary = _CATEGORY_PROMPTS.get(primary_category, {}).get("medium", "")
    secondary = _CATEGORY_PROMPTS.get(secondary_category, {}).get("medium", "")
    base = (
        f"This task could be either {HUMAN_NAMES.get(primary_category, primary_category)} "
        f"or {HUMAN_NAMES.get(secondary_category, secondary_category)}. "
        f"{primary} If it matches {HUMAN_NAMES.get(secondary_category, secondary_category)} instead: {secondary}"
    )
    return base + _ANTI_PREAMBLE


# =========================================================================
# QUALITY CONFIG
# =========================================================================

QC_CONFIG = {
    "code_debug": {"metric": "top3_gap", "threshold": 2.6667},
    "code_gen": {"metric": "margin", "threshold": 0.6},
    "factual": {"metric": "inverse_active", "threshold": 1.0},
    "logic": {"metric": "top3_gap", "threshold": 2.6667},
    "math": {"metric": "max_score", "threshold": 2.0},
    "ner": {"metric": "top_over_avg", "threshold": 4.4444},
    "sentiment": {"metric": "top3_gap", "threshold": 2.6667},
    "summarization": {"metric": "top3_gap", "threshold": 2.6667},
}
QC_POLICY = {"on_fail": "escalate", "min_accept_precision": 0.85}


# =========================================================================
# STAGE 0: PRE-FILTER
# =========================================================================

@dataclass
class Stage0Result:
    action: str = "continue"
    direct_answer: Optional[str] = None
    category: Optional[str] = None
    flags: Dict[str, bool] = field(default_factory=dict)


RE_GREETING = re.compile(r"^(hi|hello|hey|thanks|thank you|ok|okay)[\s!.]*$", re.I)
RE_MATH_OPERATION = re.compile(
    r'^[\d\s\+\-\*\/\(\)\%\.\,]+$'
)


def stage0(prompt: str) -> Stage0Result:
    """Pre-filter: bypass trivial cases."""
    cleaned = prompt.strip()

    # Tier 0A: Greetings
    if RE_GREETING.match(cleaned):
        return Stage0Result(action="bypass", direct_answer="Hello!", category="general")

    # Tier 0B: Pure math (immediate calc)
    if RE_MATH_OPERATION.match(cleaned):
        result = solve_arithmetic(cleaned)
        if result:
            return Stage0Result(action="bypass", direct_answer=result, category="math")
    return Stage0Result(action="continue")


# =========================================================================
# QUALITY VERIFICATION
# =========================================================================

_DEGENERATE_PATTERNS = [
    r"\bi don'?t know\b", r"\bi do not know\b", r"\bi cannot\b",
    r"\bi can'?t\b", r"\bas an ai\b", r"\bunable to\b",
    r"\bno information\b", r"\bis not provided\b", r"\bdoes not contain\b",
    r"\bcannot answer\b", r"\bcannot provide\b", r"\bnot enough information\b",
    r"\binsufficient\b", r"\bsorry\b", r"\bthe text does not\b",
]


def _has_hedge(text: str) -> bool:
    low = text.lower()
    for pat in _DEGENERATE_PATTERNS:
        if re.search(pat, low):
            return True
    return False


def _is_degenerate(text: str) -> bool:
    words = text.split()
    if len(words) < 2:
        return True
    # Heavy repetition
    if len(set(words)) / max(len(words), 1) < 0.3:
        return True
    return False


@dataclass
class QCResult:
    passed: bool
    reason: str = ""


def verify(text: str, task: str = "") -> QCResult:
    """Quality check on solver output."""
    text = text.strip()
    if not text:
        return QCResult(passed=False, reason="empty")
    if _has_hedge(text):
        return QCResult(passed=False, reason="hedge_detected")
    if _is_degenerate(text):
        return QCResult(passed=False, reason="degenerate")
    return QCResult(passed=True)


# =========================================================================
# LOCAL MODEL INFERENCE
# =========================================================================

_LLM: Optional[object] = None
_LLM_LOCK = threading.Lock()


def _get_llm(model_path: str, n_gpu_layers: int = -1) -> Optional[object]:
    """Lazy-load the Llama model singleton."""
    global _LLM
    if _LLM is not None:
        return _LLM
    with _LLM_LOCK:
        if _LLM is not None:
            return _LLM
        try:
            from llama_cpp import Llama
            logger.info(f"Loading model from {model_path} (n_gpu_layers={n_gpu_layers})")
            _LLM = Llama(
                model_path=model_path,
                n_ctx=N_CTX,
                n_gpu_layers=n_gpu_layers,
                n_threads=N_THREADS if n_gpu_layers == 0 else 2,
                verbose=False,
            )
            return _LLM
        except Exception as e:
            logger.error(f"Failed to load model: {e}")
            return None


def reset_model() -> None:
    """Unload the current model singleton. Must call between model swaps."""
    global _LLM
    with _LLM_LOCK:
        if _LLM is not None:
            del _LLM
            _LLM = None
            import gc
            gc.collect()
            logger.info("Model unloaded")




def solve_with_consensus(prompt: str, category: str,
                         system_prompt: str = "",
                         k: int = 1, max_tokens: int = 256,
                         model_path: str = DEFAULT_MODEL,
                         n_gpu_layers: int = -1) -> Optional[Dict[str, Any]]:
    """Single sample (k=1) inference via llama-cpp-python."""
    llm = _get_llm(model_path, n_gpu_layers)
    if llm is None:
        return None

    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    try:
        resp = llm.create_chat_completion(
            messages,
            max_tokens=max_tokens,
            temperature=0.1,
            stop=["\n\n", "<|im_end|>", "<|eot_id|>"],
        )
        answer = resp["choices"][0]["message"]["content"].strip()
        usage = resp.get("usage", {})
        logger.info(f"  {category}: {len(answer)} chars, "
                     f"prompt_tok={usage.get('prompt_tokens',0)}, "
                     f"comp_tok={usage.get('completion_tokens',0)}")
        return {
            "majority_answer": answer,
            "agreement_score": 1.0,
            "usage": usage,
            "finish_reason": resp["choices"][0].get("finish_reason", ""),
        }
    except Exception as e:
        logger.warning(f"  inference error: {e}")
        return None


# =========================================================================
# FUZZY MATCH GRADER
# =========================================================================

def _tokenize(text: str) -> List[str]:
    """Simple whitespace + punctuation tokenization."""
    return re.findall(r'\w+', text.lower())


def fuzzy_match(answer: str, expected: str) -> float:
    """Grade answer against expected using cascade: exact → substring → numeric → token overlap."""
    a = answer.strip().lower()
    e = expected.strip().lower()

    if not a or not e:
        return 0.0

    # Exact match
    if a == e:
        return 1.0

    # Substring match (expected is contained in answer or vice versa)
    if e in a or a in e:
        return 0.95

    # Numeric tolerance
    nums_a = re.findall(r'-?\d+\.?\d*', a)
    nums_e = re.findall(r'-?\d+\.?\d*', e)
    if nums_a and nums_e and len(nums_a) == len(nums_e):
        try:
            matches = sum(1 for na, ne in zip(nums_a, nums_e)
                         if abs(float(na) - float(ne)) / max(abs(float(ne)), 1) < 0.01)
            if matches == len(nums_a):
                return 0.9
        except ValueError:
            pass

    # Token overlap (Jaccard)
    tok_a = set(_tokenize(a))
    tok_e = set(_tokenize(e))
    if not tok_a or not tok_e:
        return 0.0
    overlap = len(tok_a & tok_e)
    jaccard = overlap / len(tok_a | tok_e)
    return jaccard


# =========================================================================
# MAIN PIPELINE
# =========================================================================

def run_pipeline(prompt: str, model_path: str = DEFAULT_MODEL,
                 n_gpu_layers: int = -1) -> Dict[str, Any]:
    """Run the full pipeline on one prompt. Returns dict with all results."""
    t_start = time.time()
    result = {
        "answer": "",
        "category": "",
        "complexity": 0.0,
        "used_api": False,
        "pipeline_stages": {},
        "timing_ms": {},
    }

    # ── Stage 0: Pre-filter ──
    t0 = time.time()
    s0 = stage0(prompt)
    result["timing_ms"]["stage0"] = (time.time() - t0) * 1000
    result["pipeline_stages"]["stage0"] = s0.action

    if s0.action == "bypass" and s0.direct_answer:
        result["answer"] = s0.direct_answer
        result["category"] = s0.category or "general"
        result["complexity"] = 0.0
        result["timing_ms"]["total"] = (time.time() - t_start) * 1000
        return result

    # ── Stage 2: Category classifier ──
    t1 = time.time()
    category, confidence, scores = classify(prompt)
    result["timing_ms"]["stage2"] = (time.time() - t1) * 1000
    result["category"] = category
    result["pipeline_stages"]["category"] = category
    result["pipeline_stages"]["confidence"] = confidence
    result["pipeline_stages"]["scores"] = scores

    # ── Stage 3: Complexity ──
    t2 = time.time()
    complexity = score_complexity(prompt, category)
    result["timing_ms"]["stage3"] = (time.time() - t2) * 1000
    result["complexity"] = complexity

    # ── Stage 4: Decision ──
    needs_api = True
    if complexity < COMPLEXITY_SIMPLE_MAX and category in DETERMINISTIC_CATEGORIES and category not in NAKED_CATEGORIES:
        det_cat = DET_CATEGORY_MAP.get(category, "other_complex")
        solver_fn = DET_SOLVER_FNS.get(det_cat)
        if solver_fn:
            t3 = time.time()
            try:
                ans = solver_fn(prompt, det_cat)
                result["timing_ms"]["deterministic"] = (time.time() - t3) * 1000
                if ans:
                    result["answer"] = ans
                    result["pipeline_stages"]["solver"] = f"deterministic:{det_cat}"
                    needs_api = False
            except Exception as e:
                logger.warning(f"  Deterministic solver error: {e}")

    # ── API path ──
    if needs_api:
        result["pipeline_stages"]["solver"] = "local_llm"

        # Build system prompt
        sys_prompt = ""
        if category not in NAKED_CATEGORIES:
            ner_example = NER_ONE_SHOT_EXAMPLE if category == "ner" else None
            sys_prompt = build_system_prompt(category, complexity, custom_instructions=ner_example or "")
            result["pipeline_stages"]["prompt_tier"] = (
                "low" if complexity < 0.3 else ("medium" if complexity < 0.6 else "high")
            )

        t4 = time.time()
        result_obj = solve_with_consensus(
            prompt=prompt,
            category=category,
            system_prompt=sys_prompt,
            k=1,
            max_tokens=get_max_tokens(category, complexity),
            model_path=model_path,
            n_gpu_layers=n_gpu_layers,
        )
        result["timing_ms"]["local_llm"] = (time.time() - t4) * 1000

        if result_obj and result_obj.get("majority_answer"):
            result["answer"] = result_obj["majority_answer"]
            result["pipeline_stages"]["agreement"] = result_obj.get("agreement_score", 0)
            if result_obj.get("usage"):
                result["pipeline_stages"]["prompt_tokens"] = result_obj["usage"].get("prompt_tokens", 0)
                result["pipeline_stages"]["completion_tokens"] = result_obj["usage"].get("completion_tokens", 0)

    # ── QC Gate ──
    if result.get("answer"):
        qc = verify(result["answer"], task=category)
        result["pipeline_stages"]["qc_pass"] = qc.passed
        if not qc.passed:
            result["pipeline_stages"]["qc_reason"] = qc.reason
            result["answer"] = ""  # Discard failed answer

    result["timing_ms"]["total"] = (time.time() - t_start) * 1000
    return result


def grade_answer(answer: str, expected: str) -> Dict[str, Any]:
    """Grade a single answer against expected."""
    score = fuzzy_match(answer, expected)
    return {
        "score": score,
        "passed": score >= 0.5,
        "answer": answer,
        "expected": expected,
    }


# =========================================================================
# EVAL RUNNER
# =========================================================================

def load_eval_data(path: str) -> List[Dict[str, Any]]:
    """Load eval data from JSON file (supports training-v1.json format)."""
    with open(path) as f:
        data = json.load(f)
    # Handle dict with 'questions' key (eval_clean_val.json format)
    if isinstance(data, dict) and "questions" in data:
        data = data["questions"]
    logger.info(f"Loaded {len(data)} questions from {path}")
    return data


def run_eval(eval_path: str, model_path: str = DEFAULT_MODEL,
             n_gpu_layers: int = -1, max_questions: int = 0,
             results_dir: str = "results",
             model_label: str = "") -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Run full pipeline on eval dataset. Returns per-question results."""
    os.makedirs(results_dir, exist_ok=True)
    questions = load_eval_data(eval_path)

    if max_questions > 0:
        questions = questions[:max_questions]
        logger.info(f"Limited to {max_questions} questions")

    # Warm up model
    logger.info("Warming up model...")
    t_warm = time.time()
    llm = _get_llm(model_path, n_gpu_layers)
    if llm:
        try:
            llm.create_chat_completion(
                [{"role": "user", "content": "Hello"}], max_tokens=5
            )
        except Exception:
            pass
        logger.info(f"Model warmup: {time.time() - t_warm:.1f}s")
    else:
        logger.warning("Model not loaded — pipeline will use deterministic solvers only")

    # Stats tracking
    per_category = {}
    total_time = 0.0
    results = []

    for idx, q in enumerate(questions):
        prompt = q.get("prompt", q.get("text", ""))
        expected = q.get("expected_answer", q.get("label", q.get("answer", "")))
        category = q.get("category_label", q.get("category", ""))
        task_id = q.get("task_id", q.get("id", f"q_{idx}"))

        logger.info(f"[{idx + 1}/{len(questions)}] {task_id} ({category}): {prompt[:60]}...")

        t_q = time.time()
        pipe_result = run_pipeline(prompt, model_path, n_gpu_layers)
        q_time = time.time() - t_q

        # Grade
        grading = grade_answer(pipe_result.get("answer", ""), expected)

        # Collect
        result_row = {
            "idx": idx,
            "task_id": task_id,
            "model_label": model_label,
            "category": pipe_result.get("category", category),
            "prompt": prompt[:120],
            "expected": expected[:120],
            "answer": pipe_result.get("answer", "")[:120],
            "score": round(grading["score"], 4),
            "passed": grading["passed"],
            "total_time_s": round(q_time, 3),
            "stage0_ms": round(pipe_result["timing_ms"].get("stage0", 0), 1),
            "stage2_ms": round(pipe_result["timing_ms"].get("stage2", 0), 1),
            "stage3_ms": round(pipe_result["timing_ms"].get("stage3", 0), 1),
            "det_ms": round(pipe_result["timing_ms"].get("deterministic", 0), 1),
            "llm_ms": round(pipe_result["timing_ms"].get("local_llm", 0), 1),
            "complexity": round(pipe_result.get("complexity", 0), 3),
            "confidence": round(pipe_result.get("pipeline_stages", {}).get("confidence", 0), 3),
            "raw_scores": pipe_result.get("pipeline_stages", {}).get("scores", {}),
            "max_tokens": get_max_tokens(pipe_result.get("category", category) or "factual", pipe_result.get("complexity", 0.5)),
            "qc_pass": pipe_result.get("pipeline_stages", {}).get("qc_pass", False),
            "solver": pipe_result.get("pipeline_stages", {}).get("solver", ""),
            "prompt_tokens": pipe_result.get("pipeline_stages", {}).get("prompt_tokens", 0),
            "completion_tokens": pipe_result.get("pipeline_stages", {}).get("completion_tokens", 0),
        }
        results.append(result_row)

        # Per-category stats
        cat_name = pipe_result.get("category", category) or "general"
        if cat_name not in per_category:
            per_category[cat_name] = {"count": 0, "passed": 0, "time": 0.0}
        per_category[cat_name]["count"] += 1
        per_category[cat_name]["passed"] += 1 if grading["passed"] else 0
        per_category[cat_name]["time"] += q_time
        total_time += q_time

        if (idx + 1) % 100 == 0:
            passed_sofar = sum(1 for r in results if r["passed"])
            logger.info(f"  --- [{idx+1}/{len(questions)}] accuracy so far: {passed_sofar}/{idx+1} = {passed_sofar/(idx+1)*100:.1f}%")

    # Compute summary
    summary = {
        "eval_path": eval_path,
        "model_path": model_path,
        "n_gpu_layers": n_gpu_layers,
        "total_questions": len(results),
        "total_time_s": round(total_time, 1),
        "avg_time_per_q_s": round(total_time / max(len(results), 1), 3),
        "questions_per_min": round(len(results) / max(total_time / 60, 0.01), 1),
        "overall_accuracy": round(
            sum(1 for r in results if r["passed"]) / max(len(results), 1) * 100, 1
        ),
        "passed": sum(1 for r in results if r["passed"]),
        "failed": sum(1 for r in results if not r["passed"]),
        "timestamp": datetime.now().isoformat(),
        "per_category": {},
    }
    for cat, stats in sorted(per_category.items()):
        accuracy = round(stats["passed"] / max(stats["count"], 1) * 100, 1)
        avg_time = round(stats["time"] / max(stats["count"], 1), 3)
        summary["per_category"][cat] = {
            "count": stats["count"], "passed": stats["passed"],
            "accuracy_pct": accuracy, "avg_time_s": avg_time,
        }

    return results, summary


# =========================================================================
# OUTPUT HELPERS
# =========================================================================

def _col_letter(n: int) -> str:
    """Convert 1-indexed column number to Excel column letter (A, B, ... Z, AA, AB)."""
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + n % 26) + result
        n //= 26
    return result

def write_xlsx(results: List[Dict], summary: Dict, output_path: str) -> str:
    """Write results to xlsx matching RunLogger format: Run Meta, Questions, Raw Scores sheets.
    
    Adds Score, Passed, and Expected Answer columns to the Questions sheet.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl not available — skipping xlsx")
        return ""

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Sheet 1: Run Meta ──
    meta = wb.active
    meta.title = "Run Meta"
    meta.column_dimensions["A"].width = 30
    meta.column_dimensions["B"].width = 60

    meta_rows = [
        ("Run Number", summary.get("run_number", 1)),
        ("Run Timestamp", summary.get("timestamp", "")),
        ("Pipeline Version", "colab-pipeline"),
        ("Total Questions", summary.get("total_questions", len(results))),
        ("Eval Source", os.path.basename(summary.get("eval_path", ""))),
        ("Model Path", summary.get("model_path", "")),
        ("Model Label", summary.get("model_label", "")),
        ("Fireworks Model", ""),
        ("Fireworks Key Set", "False"),
        ("N GPU Layers", str(summary.get("n_gpu_layers", -1))),
        ("N Context", "2048"),
        ("N Threads", "4"),
        ("Total Elapsed (s)", f"{summary.get('total_time_s', 0):.1f}"),
        ("Questions Logged", str(len(results))),
        ("Overall Accuracy", f"{summary.get('overall_accuracy', 0)}%"),
        ("Avg Time Per Q (s)", f"{summary.get('avg_time_per_q_s', 0):.3f}"),
    ]
    for i, (k, v) in enumerate(meta_rows, 1):
        c1 = meta.cell(row=i, column=1, value=k); c1.font = bold; c1.border = thin_border
        c2 = meta.cell(row=i, column=2, value=v); c2.border = thin_border

    # ── Sheet 2: Questions ──
    qs = wb.create_sheet("Questions")
    cols = [
        "Run", "Task ID", "Model",
        "Input Prompt", "Expected Answer", "Final Answer", "Score", "Passed",
        "Pre-Filter Action", "Pre-Filter Answer", "Pre-Filter (ms)",
        "Category", "Confidence",
        "Raw Scores",
        "Complexity", "Complexity (ms)",
        "Solver Name", "Max Tokens",
        "Prompt Version",
        "QC Pass",
        "Det (ms)",
        "Local LLM (ms)", "Prompt Tokens", "Completion Tokens", "Total Tokens",
        "Error",
        "Total (ms)",
    ]
    for ci, h in enumerate(cols, 1):
        cell = qs.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True); cell.border = thin_border

    for ri, r in enumerate(results, 2):
        vals = [
            1, r.get("task_id", ""), r.get("model_label", ""),
            r.get("prompt", ""), r.get("expected", ""), r.get("answer", ""),
            r.get("score", 0), r.get("passed", False),
            "", "", r.get("stage0_ms", 0),
            r.get("category", ""), r.get("confidence", 0),
            "",
            r.get("complexity", 0), r.get("stage3_ms", 0),
            r.get("solver", ""), r.get("max_tokens", 0),
            "",
            r.get("qc_pass", ""),
            r.get("det_ms", 0),
            r.get("llm_ms", 0), r.get("prompt_tokens", 0), r.get("completion_tokens", 0),
            r.get("prompt_tokens", 0) + r.get("completion_tokens", 0),
            r.get("error", ""),
            round(r.get("total_time_s", 0) * 1000, 1),
        ]
        for ci, v in enumerate(vals, 1):
            cell = qs.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if isinstance(v, str) and len(v) > 300:
                cell.alignment = Alignment(wrap_text=True)

    qs.auto_filter.ref = f"A1:{_col_letter(len(cols))}{len(results) + 1}"
    qs.freeze_panes = "C2"

    # ── Sheet 3: Raw Scores ──
    raw = wb.create_sheet("Raw Scores")
    raw_cols = ["Run", "Task ID", "Category", "Confidence",
                "code_debug", "code_gen", "factual", "logic",
                "math", "ner", "sentiment", "summarization"]
    for ci, h in enumerate(raw_cols, 1):
        cell = raw.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border

    for ri, r in enumerate(results, 2):
        raw_scores = r.get("raw_scores", {}) or {}
        if isinstance(raw_scores, str):
            try: raw_scores = json.loads(raw_scores)
            except: raw_scores = {}
        vals = [
            1, r.get("task_id", ""), r.get("category", ""), r.get("confidence", 0),
            round(raw_scores.get("code_debug", 0), 1),
            round(raw_scores.get("code_gen", 0), 1),
            round(raw_scores.get("factual", 0), 1),
            round(raw_scores.get("logic", 0), 1),
            round(raw_scores.get("math", 0), 1),
            round(raw_scores.get("ner", 0), 1),
            round(raw_scores.get("sentiment", 0), 1),
            round(raw_scores.get("summarization", 0), 1),
        ]
        for ci, v in enumerate(vals, 1):
            cell = raw.cell(row=ri, column=ci, value=v); cell.border = thin_border

    raw.auto_filter.ref = f"A1:L{len(results) + 1}"
    raw.freeze_panes = "C2"

    # ── Sheet 4: Per-Category Per-Model Summary (comparison) ──
    if summary.get("per_category"):
        comp = wb.create_sheet("Per-Category Summary")
        comp_cols = ["Category", "Count", "Passed", "Accuracy %", "Avg Time (s)"]
        for ci, h in enumerate(comp_cols, 1):
            cell = comp.cell(row=1, column=ci, value=h)
            cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
        for ri, (cat, s) in enumerate(sorted(summary.get("per_category", {}).items()), 2):
            vals_cat = [cat, s["count"], s["passed"], s["accuracy_pct"], s["avg_time_s"]]
            for ci, v in enumerate(vals_cat, 1):
                cell = comp.cell(row=ri, column=ci, value=v); cell.border = thin_border

    wb.save(output_path)
    logger.info(f"xlsx written to {output_path}")
    return output_path


def write_results(results: List[Dict], summary: Dict,
                  results_dir: str = "results") -> str:
    """Write results to JSON and summary text files."""
    os.makedirs(results_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.join(results_dir, f"colab_run_{timestamp}")

    # JSON
    json_path = f"{base}.json"
    with open(json_path, "w") as f:
        json.dump({"summary": summary, "results": results}, f, indent=2)
    logger.info(f"Results written to {json_path}")

    # CSV
    csv_path = f"{base}.csv"
    if results:
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=results[0].keys())
            writer.writeheader()
            writer.writerows(results)
        logger.info(f"CSV written to {csv_path}")

    # Text summary
    txt_path = f"{base}_summary.txt"
    with open(txt_path, "w") as f:
        f.write("=" * 60 + "\n")
        f.write("COLAB PIPELINE EVAL SUMMARY\n")
        f.write(f"Run: {summary['timestamp']}\n")
        f.write("=" * 60 + "\n\n")
        f.write(f"Eval file:      {summary['eval_path']}\n")
        f.write(f"Model:          {summary['model_path']}\n")
        f.write(f"GPU layers:     {summary['n_gpu_layers']}\n")
        f.write(f"Total questions: {summary['total_questions']}\n")
        f.write(f"Passed:         {summary['passed']}\n")
        f.write(f"Failed:         {summary['failed']}\n")
        f.write(f"Overall accuracy: {summary['overall_accuracy']}%\n\n")
        f.write(f"Total time:     {summary['total_time_s']}s\n")
        f.write(f"Avg per Q:      {summary['avg_time_per_q_s']}s\n")
        f.write(f"Questions/min:  {summary['questions_per_min']}\n\n")
        f.write("Per-Category Breakdown:\n")
        f.write("-" * 60 + "\n")
        f.write(f"{'Category':<16} {'Count':>6} {'Passed':>6} {'Accuracy':>10} {'Avg Time':>10}\n")
        f.write("-" * 60 + "\n")
        for cat, s in sorted(summary["per_category"].items()):
            f.write(f"{cat:<16} {s['count']:>6} {s['passed']:>6} "
                    f"{s['accuracy_pct']:>8.1f}% {s['avg_time_s']:>8.3f}s\n")
        f.write("=" * 60 + "\n")
    logger.info(f"Summary written to {txt_path}")

    return base


def print_summary(summary: Dict):
    """Print summary to stdout."""
    print()
    print("=" * 60)
    print("COLAB PIPELINE RESULTS")
    print("=" * 60)
    print(f"  Questions:    {summary['total_questions']}")
    print(f"  Passed:       {summary['passed']}")
    print(f"  Accuracy:     {summary['overall_accuracy']}%")
    print(f"  Total time:   {summary['total_time_s']:.1f}s")
    print(f"  Avg per Q:    {summary['avg_time_per_q_s']:.3f}s")
    print(f"  Q/min:        {summary['questions_per_min']:.1f}")
    print("-" * 60)
    print(f"  {'Category':<16} {'Count':>6} {'Acc%':>6} {'Avg/s':>7}")
    print("-" * 60)
    for cat, s in sorted(summary["per_category"].items()):
        print(f"  {cat:<16} {s['count']:>6} {s['accuracy_pct']:>5.1f}% {s['avg_time_s']:>6.3f}")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(description="Colab pipeline evaluator")
    parser.add_argument("--eval", required=True, help="Path to eval JSON file")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="Path to GGUF model")
    parser.add_argument("--gpu-layers", type=int, default=N_GPU_LAYERS, help="GPU layers (-1=all)")
    parser.add_argument("--max", type=int, default=0, help="Max questions (0=all)")
    parser.add_argument("--results-dir", default="results", help="Output directory")
    parser.add_argument("--model-label", default="", help="Label for this model (e.g. qwen-1.5b)")
    args = parser.parse_args()

    results, summary = run_eval(
        eval_path=args.eval,
        model_path=args.model,
        n_gpu_layers=args.gpu_layers,
        max_questions=args.max,
        results_dir=args.results_dir,
        model_label=args.model_label,
    )

    base_path = write_results(results, summary, args.results_dir)

    # Write xlsx
    xlsx_path = os.path.join(args.results_dir, f"{os.path.basename(base_path)}.xlsx")
    write_xlsx(results, summary, xlsx_path)

    print_summary(summary)
    print(f"\nResults saved: {base_path}.*")


if __name__ == "__main__":
    main()
