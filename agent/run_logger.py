"""Run Logger — comprehensive per-query pipeline logging to Excel.

Records every stage of every question: timings, classifier scores, complexity,
decision paths, solver outputs, prompt versions. Writes to structured .xlsx.

Usage:
    from agent.run_logger import RunLogger
    logger = RunLogger(run_number=1, pipeline_version="v6.1-25")
    logger.start_question("q1", "prompt text")
    logger.log_category_filter(...)
    logger.finish_question(...)
    logger.write_xlsx("/output")
"""

import json
import os
import time
from dataclasses import dataclass, field
from typing import Optional


# ═════════════════════════════════════════════════════════════════════════════
# Per-question log record
# ═════════════════════════════════════════════════════════════════════════════

@dataclass
class QuestionLog:
    """Every field captured for one question through the pipeline."""

    # ── Run metadata ──
    run_number: int = 0
    run_timestamp: str = ""

    # ── Question I/O ──
    task_id: str = ""
    model_name: str = ""            # which model answered (basename w/o path)
    input_prompt: str = ""          # truncated to first 200 chars for readability
    input_prompt_full: str = ""     # full prompt (may be long)
    final_answer: str = ""
    difficulty: str = ""            # task difficulty level from eval data

    # ── Pre-Filter (was Stage 0) ──
    pre_filter_action: str = ""     # bypass | route_to_stage3 | continue
    pre_filter_answer: str = ""
    pre_filter_category_hint: str = ""
    pre_filter_flags: str = ""
    pre_filter_ms: float = 0.0

    # ── Category Filter (was Stage 2 — 8-way classifier) ──
    category: str = ""
    category_4way: str = ""
    confidence: float = 0.0
    score_delta: float = 0.0
    raw_scores_json: str = ""       # {"code_debug": 0.0, "code_gen": 0.0, ...}
    category_filter_ms: float = 0.0
    keyword_overrides_applied: str = ""  # which overrides fired, if any

    # ── Complexity (MiniLM + LogReg) ──
    complexity: float = 0.0
    complexity_model: str = ""
    complexity_ms: float = 0.0

    # ── Decision Table output ──
    solver_name: str = ""           # deterministic | fireworks | local
    model: str = ""
    max_tokens: int = 0
    temperature: float = 0.0
    reasoning_effort: str = ""
    prefill: str = ""
    skip_api: bool = False
    decision_ms: float = 0.0

    # ── System prompt ──
    system_prompt: str = ""
    prompt_version: str = ""        # e.g. "code_gen/low", "merged/math+logic"
    is_merged: bool = False
    is_reasoning_prompt: bool = False

    # ── Deterministic solver results ──
    deterministic_hint: str = ""
    deterministic_confidence: float = 0.0
    deterministic_solver_results: str = ""  # JSON list of (solver_name, answer)
    deterministic_ms: float = 0.0

    # ── Fireworks escalation ──
    fireworks_route: str = ""       # sentiment | hard_math | hard_logic | code | empty
    fireworks_model: str = ""
    fireworks_ms: float = 0.0

    # ── Local LLM inference ──
    local_llm_ms: float = 0.0
    local_llm_retry: bool = False
    prompt_tokens: int = 0
    completion_tokens: int = 0
    total_tokens: int = 0

    # ── Post-processing ──
    post_processing_applied: str = ""  # e.g. "math_answer_extract", "logic_bare_letter_expand"

    # ── Fallback info ──
    fallback_reason: str = ""       # syntax_check_failed | local_empty |
                                    # fireworks_empty | local_timeout

    # ── Timing ──
    total_ms: float = 0.0
    solver_ms: float = 0.0

    # ── Error ──
    error: str = ""

    @property
    def classifier_ms(self) -> float:
        """Alias for category_filter_ms for backward compatibility."""
        return self.category_filter_ms

    @classifier_ms.setter
    def classifier_ms(self, value: float):
        self.category_filter_ms = value


# ═════════════════════════════════════════════════════════════════════════════
# Run counter persistence
# ═════════════════════════════════════════════════════════════════════════════

_RUN_COUNTER_PATH = os.path.join(os.path.dirname(__file__), "..", "run_counter.json")


def _load_run_number() -> int:
    try:
        if os.path.exists(_RUN_COUNTER_PATH):
            with open(_RUN_COUNTER_PATH) as f:
                return json.load(f).get("run_number", 0) + 1
    except Exception:
        pass
    return 1


def _save_run_number(run: int):
    try:
        with open(_RUN_COUNTER_PATH, "w") as f:
            json.dump({"run_number": run}, f)
    except Exception:
        pass


# ═════════════════════════════════════════════════════════════════════════════
# RunLogger
# ═════════════════════════════════════════════════════════════════════════════

class RunLogger:
    """Accumulates per-question logs and writes to Excel at end of run."""

    def __init__(
        self,
        run_number: Optional[int] = None,
        pipeline_version: str = "unknown",
        model_path: str = "",
        fireworks_model: str = "",
        fireworks_key_set: bool = False,
        n_gpu_layers: int = -1,
        n_ctx: int = 2048,
        n_threads: int = 2,
        num_questions: int = 0,
        eval_source: str = "",
    ):
        self.run_number = run_number if run_number is not None else _load_run_number()
        _save_run_number(self.run_number)

        self.pipeline_version = pipeline_version
        self.model_path = model_path
        self.fireworks_model = fireworks_model
        self.fireworks_key_set = fireworks_key_set
        self.n_gpu_layers = n_gpu_layers
        self.n_ctx = n_ctx
        self.n_threads = n_threads
        self.num_questions = num_questions
        self.eval_source = eval_source
        self.run_timestamp = time.strftime("%Y-%m-%d_%H%M%S")
        self.run_start = time.monotonic()

        self.questions: list[QuestionLog] = []
        self._current: Optional[QuestionLog] = None
        self._question_start: float = 0.0

    # ── Lifecycle ──────────────────────────────────────────────────────────

    def start_question(self, task_id: str, prompt: str, model_name: str = "",
                       difficulty: str = ""):
        """Begin logging a new question."""
        self._current = QuestionLog(
            run_number=self.run_number,
            run_timestamp=self.run_timestamp,
            task_id=task_id,
            model_name=model_name,
            input_prompt=prompt[:200],
            input_prompt_full=prompt,
            difficulty=difficulty,
        )
        self._question_start = time.monotonic()

    def finish_question(self, answer: str, solver_ms: float = 0.0, error: str = ""):
        """Finalise the current question and append to the log."""
        if self._current is None:
            return
        elapsed = (time.monotonic() - self._question_start) * 1000
        self._current.final_answer = answer
        self._current.solver_ms = solver_ms
        self._current.total_ms = elapsed
        self._current.error = error
        self.questions.append(self._current)
        self._current = None

    # ── Filter loggers ─────────────────────────────────────────────────────

    def log_pre_filter(self, action: str, answer: str, category_hint: str = "",
                       flags: str = "", elapsed_ms: float = 0.0):
        if self._current is None:
            return
        self._current.pre_filter_action = action
        self._current.pre_filter_answer = answer or ""
        self._current.pre_filter_category_hint = category_hint
        self._current.pre_filter_flags = flags
        self._current.pre_filter_ms = elapsed_ms

    def log_category_filter(self, category: str, category_4way: str, confidence: float,
                            score_delta: float, raw_scores: dict,
                            overrides: str = "", elapsed_ms: float = 0.0):
        if self._current is None:
            return
        self._current.category = category
        self._current.category_4way = category_4way
        self._current.confidence = confidence
        self._current.score_delta = score_delta
        self._current.raw_scores_json = json.dumps(raw_scores)
        self._current.keyword_overrides_applied = overrides
        self._current.category_filter_ms = elapsed_ms

    def log_complexity(self, complexity: float, model_info: str = "",
                       elapsed_ms: float = 0.0):
        if self._current is None:
            return
        self._current.complexity = complexity
        self._current.complexity_model = model_info
        self._current.complexity_ms = elapsed_ms

    def log_decision(self, solver_name: str, model: str, max_tokens: int,
                     temperature: float, reasoning_effort: str = "",
                     prefill: str = "", skip_api: bool = False,
                     system_prompt: str = "", prompt_version: str = "",
                     is_merged: bool = False, is_reasoning_prompt: bool = False,
                     elapsed_ms: float = 0.0):
        if self._current is None:
            return
        self._current.solver_name = solver_name
        self._current.model = model
        self._current.max_tokens = max_tokens
        self._current.temperature = temperature
        self._current.reasoning_effort = reasoning_effort
        self._current.prefill = prefill
        self._current.skip_api = skip_api
        self._current.system_prompt = system_prompt
        self._current.prompt_version = prompt_version
        self._current.is_merged = is_merged
        self._current.is_reasoning_prompt = is_reasoning_prompt
        self._current.decision_ms = elapsed_ms

    def log_deterministic(self, solver_results: list, hint: str = "",
                          confidence: float = 0.0, elapsed_ms: float = 0.0):
        if self._current is None:
            return
        self._current.deterministic_solver_results = json.dumps(solver_results)
        self._current.deterministic_hint = hint
        self._current.deterministic_confidence = confidence
        self._current.deterministic_ms = elapsed_ms

    def log_fireworks(self, route: str, model: str = "", elapsed_ms: float = 0.0):
        if self._current is None:
            return
        self._current.fireworks_route = route
        self._current.fireworks_model = model
        self._current.fireworks_ms = elapsed_ms

    def log_local_llm(self, elapsed_ms: float = 0.0, retry: bool = False,
                      prompt_tokens: int = 0, completion_tokens: int = 0,
                      total_tokens: int = 0):
        if self._current is None:
            return
        self._current.local_llm_ms = elapsed_ms
        self._current.local_llm_retry = retry
        self._current.prompt_tokens = prompt_tokens
        self._current.completion_tokens = completion_tokens
        self._current.total_tokens = total_tokens

    def log_post_processing(self, applied: str = ""):
        if self._current is None:
            return
        self._current.post_processing_applied = applied

    def log_fallback(self, reason: str = ""):
        if self._current is None:
            return
        self._current.fallback_reason = reason

    # ── Excel output ───────────────────────────────────────────────────────

    def write_xlsx(self, output_dir: str = "."):
        """Write all accumulated log data to a timestamped Excel workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        os.makedirs(output_dir, exist_ok=True)
        fname = f"run_{self.run_number:03d}_{self.run_timestamp}.xlsx"
        fpath = os.path.join(output_dir, fname)

        wb = openpyxl.Workbook()

        # ── Sheet 1: Run Meta ─────────────────────────────────────────────
        meta = wb.active
        meta.title = "Run Meta"
        meta.column_dimensions["A"].width = 30
        meta.column_dimensions["B"].width = 60

        bold = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        meta_rows = [
            ("Run Number", str(self.run_number)),
            ("Run Timestamp", self.run_timestamp),
            ("Pipeline Version", self.pipeline_version),
            ("Total Questions", str(self.num_questions)),
            ("Eval Source", self.eval_source),
            ("Model Path", self.model_path),
            ("Fireworks Model", self.fireworks_model),
            ("Fireworks Key Set", str(self.fireworks_key_set)),
            ("N GPU Layers", str(self.n_gpu_layers)),
            ("N Context", str(self.n_ctx)),
            ("N Threads", str(self.n_threads)),
            ("Total Elapsed (s)", f"{time.monotonic() - self.run_start:.1f}"),
            ("Questions Logged", str(len(self.questions))),
        ]
        for i, (k, v) in enumerate(meta_rows, 1):
            c1 = meta.cell(row=i, column=1, value=k)
            c2 = meta.cell(row=i, column=2, value=v)
            c1.font = bold
            c1.border = thin_border
            c2.border = thin_border

        # ── Sheet 2: Questions ────────────────────────────────────────────
        qs = wb.create_sheet("Questions")

        # Column headers
        cols = [
            "Run", "Task ID", "Model",
            "Input Prompt", "Final Answer",
            "Pre-Filter Action", "Pre-Filter Answer", "Pre-Filter (ms)",
            "Category", "Category 4-Way", "Confidence", "Score Delta",
            "Raw Scores",
            "Overrides",
            "Complexity", "Complexity Model", "Complexity (ms)",
            "Solver Name", "Model", "Max Tokens", "Temperature",
            "Reasoning Effort", "Prefill", "Skip API",
            "Prompt Version", "Is Merged", "Is Reasoning",
            "System Prompt",
            "Difficulty",
            "Det Solver Results", "Det Hint", "Det Confidence", "Det (ms)",
            "FW Route", "FW Model", "FW (ms)",
            "Local LLM (ms)", "Local LLM Retry",
            "Prompt Tokens", "Completion Tokens", "Total Tokens",
            "Post-Processing",
            "Fallback Reason",
            "Error",
            "Total (ms)", "Solver (ms)",
        ]
        qs.column_dimensions["A"].width = 5    # Run
        qs.column_dimensions["B"].width = 10   # Task ID
        qs.column_dimensions["C"].width = 60   # Input Prompt
        qs.column_dimensions["D"].width = 40   # Final Answer

        for ci, h in enumerate(cols, 1):
            cell = qs.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)
            cell.border = thin_border

        # Data rows
        for ri, q in enumerate(self.questions, 2):
            vals = [
                q.run_number, q.task_id, q.model_name,
                q.input_prompt, q.final_answer,
                q.pre_filter_action, q.pre_filter_answer, round(q.pre_filter_ms, 1),
                q.category, q.category_4way, round(q.confidence, 3),
                round(q.score_delta, 3), q.raw_scores_json,
                q.keyword_overrides_applied,
                round(q.complexity, 3), q.complexity_model,
                round(q.complexity_ms, 1),
                q.solver_name, q.model, q.max_tokens, q.temperature,
                q.reasoning_effort, q.prefill, q.skip_api,
                q.prompt_version, q.is_merged, q.is_reasoning_prompt,
                q.system_prompt,
                q.difficulty,
                q.deterministic_solver_results, q.deterministic_hint,
                round(q.deterministic_confidence, 3), round(q.deterministic_ms, 1),
                q.fireworks_route, q.fireworks_model, round(q.fireworks_ms, 1),
                round(q.local_llm_ms, 1), q.local_llm_retry,
                q.prompt_tokens, q.completion_tokens, q.total_tokens,
                q.post_processing_applied,
                q.fallback_reason,
                q.error,
                round(q.total_ms, 1), round(q.solver_ms, 1),
            ]
            for ci, v in enumerate(vals, 1):
                cell = qs.cell(row=ri, column=ci, value=v)
                cell.border = thin_border
                if isinstance(v, str) and len(v) > 300:
                    cell.alignment = Alignment(wrap_text=True)

        # Auto-filter
        qs.auto_filter.ref = f"A1:{chr(64 + len(cols))}{len(self.questions) + 1}"
        qs.freeze_panes = "C2"

        # ── Sheet 3: Raw Scores Pivot ─────────────────────────────────────
        raw = wb.create_sheet("Raw Scores")
        raw_cols = [
            "Run", "Task ID", "Category", "Confidence",
            "code_debug", "code_gen", "factual", "logic",
            "math", "ner", "sentiment", "summarization",
        ]
        for ci, h in enumerate(raw_cols, 1):
            cell = raw.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for ri, q in enumerate(self.questions, 2):
            scores = {}
            try:
                scores = json.loads(q.raw_scores_json) if q.raw_scores_json else {}
            except json.JSONDecodeError:
                pass
            vals = [
                q.run_number, q.task_id, q.category, round(q.confidence, 3),
                round(scores.get("code_debug", 0), 1),
                round(scores.get("code_gen", 0), 1),
                round(scores.get("factual", 0), 1),
                round(scores.get("logic", 0), 1),
                round(scores.get("math", 0), 1),
                round(scores.get("ner", 0), 1),
                round(scores.get("sentiment", 0), 1),
                round(scores.get("summarization", 0), 1),
            ]
            for ci, v in enumerate(vals, 1):
                cell = raw.cell(row=ri, column=ci, value=v)
                cell.border = thin_border

        raw.auto_filter.ref = f"A1:L{len(self.questions) + 1}"
        raw.freeze_panes = "C2"

        wb.save(fpath)
        return fpath
